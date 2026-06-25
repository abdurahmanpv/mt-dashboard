"""
sheet_config.py
===============
Builds the ⚙️ Config sheet — the only user-editable sheet in the workbook.

The Config sheet lets the CEO / analytics team adjust alert thresholds
without touching any Python code. Python reads these values at runtime
via data_loader.load_config() before computing decision signals.

Layout:
  Col A : Threshold Key   (read by Python — do not rename)
  Col B : Value           (edit this column to change thresholds)
  Col C : Description     (human-readable explanation)
  Col D : Default Value   (reference — what ships out of the box)
"""

import openpyxl
from styles import (C, fill, font, border, align, write_cell,
                    write_section_header, spacer, set_col_widths,
                    write_header_row)
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone


def build_config_sheet(wb: openpyxl.Workbook, existing_config: dict = None):
    """
    Create (or recreate) the ⚙️ Config sheet in the workbook.

    If existing_config is provided, the current threshold values are
    preserved (so a daily rebuild doesn't overwrite user edits).

    Parameters
    ----------
    wb              : openpyxl Workbook to add the sheet to
    existing_config : dict of {key: value} read from the current Config sheet
                      before the rebuild. Pass None to use defaults.
    """
    # Remove old sheet if exists
    sheet_name = "⚙️ Config"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Column widths: [margin, key, value, description, default, margin]
    set_col_widths(ws, [0.8, 28, 14, 52, 14, 0.8], start_col=1)

    r = 1

    # ── TITLE BANNER ──────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 6
    r += 1

    ws.merge_cells(f"B{r}:F{r}")
    cell = ws.cell(row=r, column=2, value="  ⚙️  DASHBOARD CONFIGURATION")
    cell.font      = font(bold=True, size=16, color=C["white"])
    cell.fill      = fill(C["navy"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 38
    r += 1

    ws.merge_cells(f"B{r}:F{r}")
    cell = ws.cell(row=r, column=2,
                   value="  Edit the VALUE column to adjust alert thresholds. "
                         "Do NOT rename keys in column A — Python reads them by name.")
    cell.font      = font(italic=True, size=10, color=C["white"])
    cell.fill      = fill(C["teal"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 20
    r += 1

    r = spacer(ws, r, 10)

    # ── THRESHOLD TABLE ───────────────────────────────────────────────────────
    ws.merge_cells(f"B{r}:F{r}")
    cell = ws.cell(row=r, column=2, value="  ALERT THRESHOLDS")
    cell.font      = font(bold=True, size=11, color=C["white"])
    cell.fill      = fill(C["sub_hdr"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 22
    r += 1

    # Header row
    write_header_row(ws, r,
                     ["Threshold Key", "Value", "Description", "Default"],
                     bg=C["navy"], start_col=2, height=18)
    r += 1

    # ── Threshold definitions ─────────────────────────────────────────────────
    # Each entry: (key, default_value, description)
    thresholds = [
        # CHURN
        ("churn_rate_warn",
         0.05,
         "Churn rate at or above this % triggers an AMBER (warning) alert. "
         "Industry benchmark for healthy SaaS is < 5%."),

        ("churn_rate_bad",
         0.08,
         "Churn rate at or above this % triggers a RED (critical) alert. "
         "Requires immediate retention action."),

        # TRIAL → PAID CONVERSION
        ("trial_conv_warn",
         0.30,
         "Trial→Paid conversion below this % triggers an AMBER alert. "
         "Only trials that started 7+ days ago are counted."),

        ("trial_conv_bad",
         0.20,
         "Trial→Paid conversion below this % triggers a RED alert. "
         "Investigate trial experience and upgrade prompts."),

        # FREE → PAID CONVERSION
        ("free_conv_warn",
         0.10,
         "Free→Paid conversion below this % triggers an AMBER alert. "
         "Review paywall messaging and upgrade CTAs."),
    ]

    for i, (key, default, description) in enumerate(thresholds):
        # Use existing config value if available (preserve user edits on rebuild)
        current_value = (existing_config or {}).get(key, default)

        bg = C["light_bg"] if i % 2 == 0 else C["white"]
        ws.row_dimensions[r].height = 36

        # Col B: key (read-only reference — styled to indicate it's a code key)
        cell = ws.cell(row=r, column=2, value=key)
        cell.font      = font(bold=True, size=10, color=C["navy"],
                              name="Courier New")
        cell.fill      = fill(bg)
        cell.border    = border("thin")
        cell.alignment = align(h="left", v="center")

        # Col C: value (editable — highlighted with gold border)
        cell = ws.cell(row=r, column=3, value=current_value)
        cell.font      = font(bold=True, size=11, color=C["text_dk"])
        cell.fill      = fill(C["warn_bg"])
        cell.border    = border("medium", C["gold"])
        cell.alignment = align(h="center", v="center")
        cell.number_format = "0.00%"  # Format as percentage

        # Col D: description
        cell = ws.cell(row=r, column=4, value=description)
        cell.font      = font(size=10, color=C["text_md"])
        cell.fill      = fill(bg)
        cell.border    = border("thin")
        cell.alignment = align(h="left", v="center", wrap=True)

        # Col E: default value (reference only)
        cell = ws.cell(row=r, column=5, value=default)
        cell.font      = font(italic=True, size=10, color=C["text_lt"])
        cell.fill      = fill(bg)
        cell.border    = border("thin")
        cell.alignment = align(h="center", v="center")
        cell.number_format = "0.00%"

        r += 1

    r = spacer(ws, r, 10)

    # ── USAGE NOTES ───────────────────────────────────────────────────────────
    ws.merge_cells(f"B{r}:F{r}")
    cell = ws.cell(row=r, column=2, value="  HOW TO USE THIS SHEET")
    cell.font      = font(bold=True, size=11, color=C["white"])
    cell.fill      = fill(C["sub_hdr"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 22
    r += 1

    notes = [
        ("Editing thresholds",
         "Click any cell in the VALUE column (Column C) and type the new threshold. "
         "Values are percentages — enter as decimals (e.g. 0.05 for 5%). "
         "Changes take effect on the next daily refresh."),
        ("Do NOT rename keys",
         "Column A contains the exact key names Python reads. Renaming them will "
         "cause the dashboard to fall back to default values silently."),
        ("Adding new thresholds",
         "To add a new threshold: add a row here with a key and value, then add "
         "the corresponding signal logic in kpi_engine.py → compute_signals()."),
        ("Rebuild without DB",
         "To rebuild the dashboard without fetching fresh MySQL data: "
         "python daily_refresh.py --skip-db"),
        ("Last rebuilt",
         f"This config sheet was last generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC"),
    ]

    for i, (label, text) in enumerate(notes):
        bg = C["card_bg"] if i % 2 == 0 else C["white"]
        ws.row_dimensions[r].height = 44

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
        cell = ws.cell(row=r, column=2, value=label)
        cell.font      = font(bold=True, size=10, color=C["navy"])
        cell.fill      = fill(bg)
        cell.border    = border("thin")
        cell.alignment = align(h="left", v="center", wrap=True)

        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        cell = ws.cell(row=r, column=3, value=text)
        cell.font      = font(size=10, color=C["text_md"])
        cell.fill      = fill(bg)
        cell.border    = border("thin")
        cell.alignment = align(h="left", v="center", wrap=True)

        r += 1

    # ── FOOTER ────────────────────────────────────────────────────────────────
    r = spacer(ws, r, 8)
    ws.merge_cells(f"B{r}:F{r}")
    cell = ws.cell(row=r, column=2,
                   value=f"  CEO Subscription Dashboard  ·  Config Sheet  ·  "
                         f"Auto-managed by build_dashboard.py  ·  "
                         f"User edits in Column C are preserved across rebuilds")
    cell.font      = font(italic=True, size=9, color=C["white"])
    cell.fill      = fill(C["navy"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 18
