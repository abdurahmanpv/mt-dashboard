import openpyxl

path = "CEO_Subscription_Dashboard.xlsx"
sheet_name = "Monthly Active Snapshots"

wb = openpyxl.load_workbook(path)

if sheet_name not in wb.sheetnames:
    print(f"Sheet '{sheet_name}' not found — nothing to clear")
else:
    ws = wb[sheet_name]
    # Keep header row (row 1), delete all data rows
    rows_before = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    wb.save(path)
    print(f"Cleared {rows_before - 1} data rows from '{sheet_name}'. Header preserved.")

wb.close()
