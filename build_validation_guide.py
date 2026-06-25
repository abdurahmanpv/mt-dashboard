"""
build_validation_guide.py
=========================
Generates CEO_Dashboard_Validation_Guide.xlsx alongside the main dashboard.
Called automatically by build_dashboard.py on every daily refresh.

The guide is a static reference document explaining how to validate every
dashboard metric from the Raw Data sheet.  It is rebuilt on each run so
the output path is always current and the file is never stale.
"""

import os
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger(__name__)

# ── Palette (matches main dashboard) ──────────────────────────────────────────
NAVY       = "1F3864"
TEAL       = "17375E"
WHITE      = "FFFFFF"
LIGHT_BLUE = "DCE6F1"
YELLOW     = "FFFF99"
GRAY       = "F2F2F2"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _hdr(ws, row, col, text, bg=NAVY, fg=WHITE, bold=True, wrap=False, size=10):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=bold, color=fg, size=size, name="Arial")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    return c


def _val(ws, row, col, text, bg=None, bold=False, wrap=True, size=10):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=bold, color="000000", size=size, name="Arial")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    return c


def _col_widths(ws, widths):
    """widths: list of (col_letter, width) tuples."""
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _sheet_how_to_use(wb):
    ws = wb.active
    ws.title = "How to Use"
    _col_widths(ws, [2, 28, 55, 2])

    _hdr(ws, 1, 2, "CEO SUBSCRIPTION DASHBOARD — VALIDATION GUIDE", size=13)
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 24

    _val(ws, 2, 2,
         "Purpose: Step-by-step instructions for validating every metric "
         "in the dashboard against the MySQL source data.",
         wrap=True)
    ws.merge_cells("B2:C2")
    ws.row_dimensions[2].height = 32

    _hdr(ws, 4, 2, "DATA SOURCE (Raw Data sheet)", bg=TEAL, size=11)
    ws.merge_cells("B4:C4")

    source_rows = [
        ("Raw Data sheet",
         "The 'Raw Data' sheet contains every order row from MySQL (all statuses). "
         "One row per billing event. QA test subscriptions are visible here but are "
         "EXCLUDED from all KPI sheets (identified by refund_reason containing 'test')."),
        ("Primary grain",
         "order_identifier — unique per billing event. A subscription can have many "
         "orders (trial → paid → renewals)."),
        ("Key filter",
         "For subscription counts: filter order_status = 'Confirmed' OR 'Completed'. "
         "Cancelled/Refund rows exist but are excluded from KPI counts."),
        ("curr_status",
         "Active = paid_till > yesterday (strictly greater, Pacific). In-Active = expired. "
         "Case B (end_date): when end_date IS NOT NULL and paid_till <= end_date, "
         "access ends at end_date instead of paid_till."),
        ("paid_till / end_date",
         "paid_till and end_date ARE present in the Raw Data sheet. "
         "paid_till = when paid access expires (Case A subscriptions). "
         "end_date = explicit termination date (Case B subscriptions — see curr_status note above). "
         "Use paid_till > yesterday for active checks; prefer curr_status for simplicity."),
        ("QA test exclusion",
         "Subscriptions where any order has refund_reason containing 'test' (case-insensitive) "
         "are excluded from ALL KPI metrics and counts. They appear in Raw Data only. "
         "If your manual count includes these, subtract them to match the dashboard."),
        ("Workbook sheets",
         "Executive Dashboard · Plan Performance · Retention & Renewals · "
         "Active Subscriptions (month-on-month active by plan) · Raw Data · Monthly Active Snapshots."),
        ("Snapshots",
         "The 'Monthly Active Snapshots' sheet captures subscription state on the "
         "1st of each month. Use it for point-in-time month-start denominators. "
         "Note: YTD and Lifetime churn/conversion rates use the AVERAGE of month-start "
         "active counts across all months in the period, not a single snapshot."),
    ]
    for i, (label, desc) in enumerate(source_rows, start=5):
        _hdr(ws, i, 2, label, bg=LIGHT_BLUE, fg="000000", bold=True)
        _val(ws, i, 3, desc, wrap=True)
        ws.row_dimensions[i].height = 40

    _hdr(ws, 12, 2, "VALIDATION STEPS (applies to all metrics)", bg=TEAL, size=11)
    ws.merge_cells("B12:C12")

    steps = [
        ("Step 1", "Open the dashboard Excel → Raw Data sheet."),
        ("Step 2", "Apply filters per the 'Metric Formulas' sheet for the metric you want to validate."),
        ("Step 3", "Compare count / sum to the corresponding cell in the dashboard."),
        ("Step 4",
         "If numbers differ, check: (a) date window matches, (b) order_status filter applied, "
         "(c) plan_segment filter correct."),
        ("Note",
         "Small residual differences (<5) can occur because the dashboard uses "
         "Pacific-timezone date arithmetic at query time; the Raw Data sheet is a static snapshot."),
    ]
    for i, (label, desc) in enumerate(steps, start=13):
        _hdr(ws, i, 2, label, bg=GRAY, fg="000000", bold=True)
        _val(ws, i, 3, desc, wrap=True)
        ws.row_dimensions[i].height = 32


def _sheet_metric_formulas(wb):
    ws = wb.create_sheet("Metric Formulas")
    _col_widths(ws, [2, 30, 22, 55, 2])

    _hdr(ws, 1, 2,
         "METRIC FORMULAS — HOW EACH DASHBOARD NUMBER IS COMPUTED",
         size=12)
    ws.merge_cells("B1:D1")
    ws.row_dimensions[1].height = 22

    _hdr(ws, 2, 2, "Metric",                          bg=TEAL)
    _hdr(ws, 2, 3, "Dashboard Sheet",                 bg=TEAL)
    _hdr(ws, 2, 4, "How to validate from Raw Data sheet", bg=TEAL)
    ws.row_dimensions[2].height = 18

    metrics = [
        (
            "Active Subscribers (Now)",
            "Executive Dashboard",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter curr_status = 'Active'\n"
            "3. Count UNIQUE subscription_id values\n\n"
            "Active = paid_till > yesterday (strictly greater, Pacific). "
            "Case B: when end_date IS NOT NULL and end_date > yesterday, "
            "the subscription is also active regardless of paid_till. "
            "curr_status = 'Active' already encodes both cases."
        ),
        (
            "Active Subscribers — Paid breakdown",
            "Executive Dashboard",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter curr_status = 'Active'\n"
            "3. Filter trial_check = 'PAID'\n"
            "4. Count UNIQUE subscription_id values"
        ),
        (
            "Active Subscribers — Free breakdown",
            "Executive Dashboard",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter curr_status = 'Active'\n"
            "3. Filter trial_check = 'FREE'\n"
            "4. Count UNIQUE subscription_id values"
        ),
        (
            "Trial Pipeline (open trials)",
            "Executive Dashboard",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter subscription_type = 'Trial'\n"
            "3. Filter curr_status = 'Active'\n"
            "4. Exclude subscription_ids that also have a trial_check = 'PAID' row\n"
            "5. Count UNIQUE subscription_id values\n\n"
            "Step 4 excludes already-converted trials — only open/unconverted "
            "trials are counted."
        ),
        (
            "New Subscriptions MTD",
            "Executive Dashboard / Plan Performance",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter subscription_type = 'New Subscription'\n"
            "3. Filter subscribed_date >= 1st of current month AND <= yesterday\n"
            "4. Count UNIQUE subscription_id values"
        ),
        (
            "Total Transactions MTD",
            "Executive Dashboard",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter subscribed_date >= 1st of current month AND <= yesterday\n"
            "3. Count ROWS (each row = one billing event)\n\n"
            "Includes trials, new subscriptions, and renewals. "
            "Previously labelled 'Billing Events'."
        ),
        (
            "Renewals MTD",
            "Executive Dashboard / Plan Performance",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter subscription_type = 'Renewal'\n"
            "3. Filter subscribed_date >= 1st of current month AND <= yesterday\n"
            "4. Count ROWS (one renewal = one order row)"
        ),
        (
            "MRR (Monthly Recurring Revenue)",
            "Executive Dashboard",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter curr_status = 'Active'\n"
            "3. Sum mrr_amount column\n\n"
            "Monthly plans: mrr_amount = total_amt. "
            "Yearly plans: mrr_amount = total_amt ÷ 12 (recognition spread)."
        ),
        (
            "Cash Revenue MTD",
            "Executive Dashboard",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter trial_check = 'PAID'\n"
            "3. Filter subscribed_date >= 1st of current month AND <= yesterday\n"
            "4. Sum total_amt\n\n"
            "For Net Cash: subtract SUM(refund_amount) for Cancelled/Refund rows "
            "where refund_date falls in the same window (refunds are bucketed by "
            "refund_date, not the original order date)."
        ),
        (
            "Paid Churn Events MTD / Prev / Yesterday",
            "Retention & Renewals",
            "1. Filter trial_check = 'PAID' (subscriptions with at least one paid order)\n"
            "2. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "3. Per subscription_id, take the MAX paid_till date\n"
            "   (Case B: use end_date if end_date IS NOT NULL and paid_till <= end_date)\n"
            "4. Filter: max paid_till falls WITHIN the window AND curr_status = 'In-Active'\n"
            "5. Count UNIQUE subscription_id values\n\n"
            "Churn date = when paid access expired, not when the order was placed. "
            "This correctly captures subscriptions that churned in the period."
        ),
        (
            "Paid Churn Rate MTD / Prev",
            "Retention & Renewals",
            "= Paid Churn Events ÷ Active Paid Subscriptions at Period Start\n\n"
            "Denominator (MTD): Monthly Active Snapshots — filter snapshot_date = "
            "1st of current month, plan_segment != 'Free Monthly', "
            "count unique subscription_id values.\n\n"
            "Denominator (Prev Full Month): same but snapshot_date = 1st of prev month."
        ),
        (
            "Paid Churn Rate YTD / Lifetime",
            "Retention & Renewals",
            "= Paid Churn Events in period ÷ AVG active paid subs at each month start\n\n"
            "Denominator = average of the active-paid count at the 1st of each month "
            "that falls within the period (Jan 1, Feb 1, … for YTD). "
            "Cannot be reproduced from a single snapshot row — requires averaging "
            "across the relevant snapshot_date rows in the Snapshots sheet."
        ),
        (
            "Trial → Paid Conversion MTD",
            "Retention & Renewals",
            "Denominator: trial subscriptions whose trial END DATE falls in MTD window.\n"
            "  - For converted trials: trial_end = first paid order date on same subscription_id\n"
            "  - For unconverted trials: trial_end = max paid_till of the trial orders\n"
            "    (falls back to subscribed_date + 7 days when paid_till is NULL)\n\n"
            "Numerator: from that denominator set, subscriptions that also placed a "
            "paid order (trial_check = 'PAID') within the same MTD window.\n\n"
            "Note: trial duration is derived from data — not hardcoded to 7 days."
        ),
        (
            "Trial → Paid Conversion YTD / Lifetime",
            "Retention & Renewals",
            "Same logic as MTD but the window is Jan 1 → yesterday (YTD) "
            "or all-time → yesterday (Lifetime). "
            "Denominator is NOT averaged — it is the count of trials whose end date "
            "falls in the period (event-based, not cohort-based)."
        ),
        (
            "Free → Paid Conversion rate MTD",
            "Retention & Renewals",
            "Numerator: user_ids who had a FREE order on any subscription AND placed "
            "a PAID order within the MTD window.\n\n"
            "Denominator (Active Free Users): free-plan user_ids whose free subscription "
            "was active at the start of the month (free paid_till >= MTD start).\n"
            "From Raw Data: filter trial_check = 'FREE', paid_till >= 1st of current month, "
            "count unique user_id values."
        ),
        (
            "Free → Paid Conversion YTD / Lifetime",
            "Retention & Renewals",
            "Same numerator logic as MTD but the window is Jan 1 → yesterday (YTD) "
            "or all-time (Lifetime).\n\n"
            "Denominator = AVERAGE of active-free-user counts at each month-start in "
            "the period (not a single snapshot). Cannot be reproduced from a single "
            "filter pass — requires averaging across multiple month-start snapshots."
        ),
        (
            "Renewals (Retention Rate MTD)",
            "Retention & Renewals",
            "Numerator:\n"
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter subscription_type = 'Renewal'\n"
            "3. Filter subscribed_date >= 1st of current month AND <= yesterday\n"
            "4. Count UNIQUE subscription_id values\n\n"
            "Denominator: paid subscriptions active at start of month.\n"
            "(Monthly Active Snapshots — filter snapshot_date = 1st of current month, "
            "plan_segment != 'Free Monthly'.)"
        ),
        (
            "Active Subscriptions by Plan (monthly)",
            "Active Subscriptions",
            "For each month-start date D (Jan 1, Feb 1, … Jun 1):\n"
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter to orders where subscribed_date <= D - 1 day\n"
            "   (exclude orders placed after D — prevents future renewals from\n"
            "    inflating counts for subscriptions that had lapsed by date D)\n"
            "3. Per subscription_id, take max paid_till of eligible orders\n"
            "   (apply Case B: use end_date when end_date IS NOT NULL and paid_till <= end_date)\n"
            "4. Filter: max paid_till > (D - 1 day)  →  active at start of D\n"
            "5. Group by plan_segment and count unique subscription_id values\n\n"
            "Trials column: trial subs (subscription_type = 'Trial') still active at D "
            "that have NOT placed a paid order before D.\n\n"
            "Note: step 2 is essential for correctness when a subscription lapsed "
            "and later re-subscribed — without it, the re-subscription's paid_till "
            "would make the sub appear continuously active through the gap period."
        ),
        (
            "Platform breakdown",
            "Plan Performance",
            "1. Filter order_status = 'Confirmed' OR 'Completed'\n"
            "2. Filter subscription_type = 'New Subscription'\n"
            "3. Filter subscribed_date in window of interest\n"
            "4. Group by platform or device_type\n"
            "5. Count UNIQUE subscription_id values per group"
        ),
    ]

    for i, (metric, sheet, steps) in enumerate(metrics, start=3):
        bg = LIGHT_BLUE if i % 2 == 1 else WHITE
        _hdr(ws, i, 2, metric,  bg=bg, fg="000000", bold=True)
        _val(ws, i, 3, sheet,   bg=bg, wrap=True)
        _val(ws, i, 4, steps,   bg=bg, wrap=True)
        ws.row_dimensions[i].height = max(60, steps.count("\n") * 16 + 20)


def _sheet_snapshots(wb):
    ws = wb.create_sheet("Metrics Requiring Snapshots")
    _col_widths(ws, [2, 34, 58, 2])

    _hdr(ws, 1, 2,
         "METRICS THAT USE THE MONTHLY ACTIVE SNAPSHOTS SHEET",
         size=12)
    ws.merge_cells("B1:C1")
    ws.row_dimensions[1].height = 22

    _val(ws, 2, 2,
         "Some metrics require a point-in-time count of active subscriptions at "
         "the START of a month. The 'Monthly Active Snapshots' sheet captures this "
         "on the 1st of each month (idempotent — one snapshot per month, written on "
         "the first daily run of that month).\n\n"
         "IMPORTANT: YTD and Lifetime churn/conversion RATES use the AVERAGE of "
         "active counts at each month-start in the period — not a single snapshot. "
         "To reproduce: sum active counts across all relevant snapshot_date rows, "
         "then divide by the number of months.",
         wrap=True, bg=YELLOW)
    ws.merge_cells("B2:C2")
    ws.row_dimensions[2].height = 72

    _hdr(ws, 3, 2, "Snapshot sheet columns", bg=TEAL)
    _hdr(ws, 3, 3,
         "snapshot_date | subscription_id | user_id | plan_segment | paid_till",
         bg=TEAL)
    ws.row_dimensions[3].height = 18

    _hdr(ws, 4, 2, "Metric",                                          bg=TEAL)
    _hdr(ws, 4, 3, "How to get denominator from Snapshots sheet",     bg=TEAL)
    ws.row_dimensions[4].height = 18

    snap_rows = [
        ("Paid Churn Rate MTD — denominator",
         "Filter snapshot_date = 1st of current month.\n"
         "Filter plan_segment != 'Free Monthly' (i.e. paid plans only).\n"
         "Count UNIQUE subscription_id values."),
        ("Paid Churn Rate Prev Full Month — denominator",
         "Filter snapshot_date = 1st of previous month.\n"
         "Filter plan_segment != 'Free Monthly'.\n"
         "Count UNIQUE subscription_id values."),
        ("Paid Churn Rate YTD — denominator (avg)",
         "For each month-start from Jan 1 through MTD_START:\n"
         "  Filter snapshot_date = that month's 1st, plan_segment != 'Free Monthly'.\n"
         "  Record the count.\n"
         "Denominator = AVERAGE of those counts across all months in the period."),
        ("Paid Churn Rate Lifetime — denominator (avg)",
         "Same as YTD but include ALL available snapshot_date values.\n"
         "Denominator = AVERAGE of active-paid counts at each month-start on record."),
        ("Free → Paid Conversion MTD — denominator",
         "Filter snapshot_date = 1st of current month.\n"
         "Filter plan_segment = 'Free Monthly'.\n"
         "Count UNIQUE user_id values (free→paid tracks users, not subscriptions)."),
        ("Free → Paid Conversion YTD / Lifetime — denominator (avg)",
         "Same avg approach as YTD churn: average of active-free user counts "
         "at each month-start in the period."),
        ("Retention Rate MTD — denominator",
         "Same as Paid Churn Rate MTD denominator.\n"
         "(Active paid subscriptions at start of month = subscriptions eligible to renew.)"),
        ("Per-segment churn denominators",
         "Filter snapshot_date = 1st of the relevant month.\n"
         "Filter plan_segment = <segment of interest>.\n"
         "Count UNIQUE subscription_id values."),
        ("Active Subscriptions sheet — alternative source",
         "The 'Active Subscriptions' sheet shows active counts per plan per month-start, "
         "computed directly from the full subscription history (not just from snapshots). "
         "It provides the same month-start counts but also covers months before the "
         "pipeline first ran."),
    ]
    for i, (metric, desc) in enumerate(snap_rows, start=5):
        bg = LIGHT_BLUE if i % 2 == 1 else WHITE
        _hdr(ws, i, 2, metric, bg=bg, fg="000000", bold=True)
        _val(ws, i, 3, desc,   bg=bg, wrap=True)
        ws.row_dimensions[i].height = max(50, desc.count("\n") * 16 + 20)


def _sheet_date_windows(wb):
    ws = wb.create_sheet("Date Windows Reference")
    _col_widths(ws, [2, 24, 32, 44, 2])

    _hdr(ws, 1, 2, "DATE WINDOWS — WHAT EACH PERIOD LABEL MEANS", size=12)
    ws.merge_cells("B1:D1")
    ws.row_dimensions[1].height = 22

    _hdr(ws, 2, 2, "Period Label",  bg=TEAL)
    _hdr(ws, 2, 3, "Date Range",    bg=TEAL)
    _hdr(ws, 2, 4, "Notes",         bg=TEAL)
    ws.row_dimensions[2].height = 18

    windows = [
        ("Yesterday",
         "Yesterday only (Pacific)",
         "subscribed_date = yesterday. Single-day window. "
         "Used for churn events, conversion counts, and cancellation counts. "
         "Rate metrics (churn rate, conv rate) for this window should be read "
         "as a daily snapshot, not a trend."),
        ("MTD (Month-to-Date)",
         "1st of current month → yesterday (Pacific)",
         "subscribed_date >= 1st of current month AND <= yesterday. "
         "All times in Pacific timezone."),
        ("Prev MTD (same-period last month)",
         "1st of prev month → same day-of-month in prev month",
         "e.g. if today is Jun 22, prev MTD = May 1 – May 21. "
         "Used for apples-to-apples period comparison in KPI cards."),
        ("Prev Full Month",
         "1st of prev month → last day of prev month",
         "Full calendar month. Used as denominator reference for single-month "
         "churn and conversion rates."),
        ("YTD (Year-to-Date)",
         "Jan 1 of current year → yesterday (Pacific)",
         "subscribed_date >= Jan 1 AND <= yesterday. "
         "Churn/conversion RATES use avg active counts at each month-start "
         "in the period as denominator."),
        ("Lifetime",
         "Earliest data → yesterday (Pacific)",
         "No start date filter. subscribed_date <= yesterday. "
         "Churn/conversion RATES use avg active counts at each month-start "
         "across all available history as denominator."),
        ("Point-in-time (Active Now)",
         "As of end of yesterday (Pacific)",
         "Not a range — a snapshot. "
         "Active definition (Case A): paid_till > yesterday (strictly greater). "
         "Active definition (Case B): when end_date IS NOT NULL and paid_till <= end_date, "
         "active if end_date > yesterday. "
         "curr_status = 'Active' in Raw Data encodes both cases."),
    ]
    for i, (label, dates, notes) in enumerate(windows, start=3):
        bg = LIGHT_BLUE if i % 2 == 1 else WHITE
        _hdr(ws, i, 2, label, bg=bg, fg="000000", bold=True)
        _val(ws, i, 3, dates, bg=bg, wrap=True)
        _val(ws, i, 4, notes, bg=bg, wrap=True)
        ws.row_dimensions[i].height = 38


# ── Public entry point ─────────────────────────────────────────────────────────

def build_validation_guide(dashboard_output_path: str) -> str:
    """
    Build (or rebuild) CEO_Dashboard_Validation_Guide.xlsx next to the
    dashboard output file.

    Parameters
    ----------
    dashboard_output_path : str
        Path to the main dashboard Excel file.
        The guide is written to the same directory with the fixed name
        CEO_Dashboard_Validation_Guide.xlsx.

    Returns
    -------
    str : path to the saved guide file
    """
    out_dir  = Path(dashboard_output_path).parent
    out_path = out_dir / "CEO_Dashboard_Validation_Guide.xlsx"

    wb = openpyxl.Workbook()
    _sheet_how_to_use(wb)
    _sheet_metric_formulas(wb)
    _sheet_snapshots(wb)
    _sheet_date_windows(wb)

    wb.save(str(out_path))
    log.info(f"  Validation guide written → {out_path}")
    return str(out_path)
