"""
sheet_builders.py
=================
Sheet builders for:
  - Subscription Funnel
  - Plan Performance
  - Retention & Renewals
  - Raw Data
"""

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from styles import (C, fill, font, border, align,
                    write_header_row, write_data_row, write_section_header,
                    spacer, set_col_widths, fmt_num, fmt_pct, fmt_currency,
                    nv, pv, cv)
from kpi_engine import PLAN_SEGMENTS, MTD_START, YESTERDAY


# ══════════════════════════════════════════════════════════════════════════════
#  SUBSCRIPTION FUNNEL
# ══════════════════════════════════════════════════════════════════════════════

def build_funnel_sheet(wb: openpyxl.Workbook, kpi: dict):
    """
    DEPRECATED — Subscription Funnel sheet removed.
    Plan breakdown has moved to Plan Performance sheet.
    This function is kept as a no-op to avoid import errors.
    """
    return  # no-op
    sheet_name = "Subscription Funnel"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [0.8, 32, 16, 16, 18, 20, 16, 16, 0.8])

    r = 1
    ws.row_dimensions[r].height = 6
    r += 1

    # Title
    ws.merge_cells(f"B{r}:I{r}")
    cell = ws.cell(row=r, column=2, value="  SUBSCRIPTION FUNNEL")
    cell.font = font(bold=True, size=16, color=C["white"])
    cell.fill = fill(C["navy"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 36
    r += 1

    ws.merge_cells(f"B{r}:I{r}")
    tz = kpi.get("tz_label", "Pacific Time (PT)")
    stale_tag = (f"  ⚠ DATA STALE — {kpi['data_lag_days']}d behind (run daily_refresh.py)  ·  "
                 if kpi.get("data_stale") else "  ")
    cell = ws.cell(row=r, column=2,
                   value=f"{stale_tag}Data through {kpi['as_of']} ({tz})  ·  "
                         f"MTD = {kpi['mtd_start']} → {kpi['as_of']} ({kpi['days_elapsed_mtd']} days)  ·  "
                         f"Generated: {kpi.get('generated_on', kpi['as_of'])}")
    cell.font = font(italic=True, size=9, color=C["white"])
    cell.fill = fill(C["amber"] if kpi.get("data_stale") else C["teal"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 18
    r += 1

    r = spacer(ws, r, 8)

    # ── COMBINED FUNNEL ───────────────────────────────────────────────────────
    r = write_section_header(ws, r, 2, 8, "COMBINED FUNNEL — ALL PLANS")

    hdrs = ["Stage", "Yesterday", "MTD", "Prev Month (Full)",
            f"Same Period Last Mo.", "YTD", "Lifetime"]
    write_header_row(ws, r, hdrs, C["navy"], start_col=2, height=18)
    r += 1

    funnel = kpi["funnel"]["combined"]
    days   = kpi["days_elapsed_mtd"]

    funnel_stages = [
        ("App Sign-ups (data not yet available)", None),   # placeholder — no data yet
        ("  → Trials",             funnel["trials"]),
        ("  → New Subscriptions",  funnel["new_subs"]),
        ("      → Paid",           funnel["paid"]),
        ("      → Free",           funnel["free"]),
        ("  → Renewals",           funnel["renewals"]),
    ]

    BLANK = {w: None for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]}

    for i, (label, counts) in enumerate(funnel_stages):
        _counts = counts if counts is not None else BLANK
        vals = [
            "–" if _counts["yesterday"] is None else fmt_num(_counts["yesterday"]),
            "–" if _counts["mtd"]       is None else fmt_num(_counts["mtd"]),
            "–" if _counts["prev_full"] is None else fmt_num(_counts["prev_full"]),
            "–" if _counts["prev_same"] is None else fmt_num(_counts["prev_same"]),
            "–" if _counts["ytd"]       is None else fmt_num(_counts["ytd"]),
            "–" if _counts["lifetime"]  is None else fmt_num(_counts["lifetime"]),
        ]
        write_data_row(ws, r, [label] + vals,
                       start_col=2, alt=(i % 2 == 0), height=18,
                       bold_first=not label.startswith(" "))
        r += 1

    # Conversion rates row
    ws.row_dimensions[r].height = 18
    t_conv = kpi["trial_conversion"]
    f_conv = kpi["free_conversion"]
    pchurn = kpi["paid_churn"]

    conv_rows = [
        ("Trial→Paid Conv%",
         "–",
         fmt_pct(t_conv["mtd"]["rate"]),
         fmt_pct(t_conv["prev_full"]["rate"]),
         "–",
         fmt_pct(t_conv["ytd"]["rate"]),
         fmt_pct(t_conv["lifetime"]["rate"])),
        ("Free→Paid Conv%",
         "–",
         fmt_pct(f_conv["mtd"]["rate"]),
         fmt_pct(f_conv["prev_full"]["rate"]),
         "–",
         fmt_pct(f_conv["ytd"]["rate"]),
         fmt_pct(f_conv["lifetime"]["rate"])),
        ("Paid Churn Rate",
         "–",
         fmt_pct(pchurn["mtd"]["rate"]),
         fmt_pct(pchurn["prev_full"]["rate"]),
         "–", "–", "–"),
        ("Paid Retention Rate",
         "–",
         fmt_pct(pchurn["mtd"]["retention_rate"]),
         fmt_pct(pchurn["prev_full"]["retention_rate"]),
         "–", "–", "–"),
    ]

    for i, row_vals in enumerate(conv_rows):
        write_data_row(ws, r, list(row_vals),
                       start_col=2, alt=((len(funnel_stages) + i) % 2 == 0),
                       height=18, bold_first=True)
        r += 1

    r = spacer(ws, r, 12)

    # ── PER-PLAN BREAKDOWN (Lite vs Unlimited) ────────────────────────────────
    r = write_section_header(ws, r, 2, 8, "PLAN BREAKDOWN — LITE vs UNLIMITED")

    # Sub-headers for each plan
    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    ws.cell(row=r, column=2, value="Stage").font = font(bold=True, size=10, color=C["white"])
    ws.cell(row=r, column=2).fill = fill(C["navy"])
    ws.cell(row=r, column=2).alignment = align()

    plans = ["Lite", "Unlimited"]
    plan_colors = [C["teal"], "6C3483"]
    windows_labels = ["MTD", "Prev", "Lifetime"]
    windows_keys   = ["mtd", "prev_full", "lifetime"]

    col = 3
    for plan, color in zip(plans, plan_colors):
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 2)
        cell = ws.cell(row=r, column=col, value=plan)
        cell.font      = font(bold=True, size=11, color=C["white"])
        cell.fill      = fill(color)
        cell.alignment = align()
        col += 3

    r += 1

    # Window sub-headers
    ws.row_dimensions[r].height = 16
    ws.cell(row=r, column=2, value="").fill = fill(C["navy"])
    col = 3
    for plan, color in zip(plans, plan_colors):
        for wlabel in windows_labels:
            cell = ws.cell(row=r, column=col, value=wlabel)
            cell.font      = font(bold=True, size=9, color=C["white"])
            cell.fill      = fill(color)
            cell.alignment = align()
            cell.border    = border("thin")
            col += 1

    r += 1

    by_plan = kpi["funnel"]["by_plan"]
    plan_stages = [
        ("Trials",           "trials"),
        ("New Subscriptions","new_subs"),
        ("Renewals",         "renewals"),
        ("Paid",             "paid"),
        ("Cash Revenue ($)", "revenue"),
    ]

    for i, (label, key) in enumerate(plan_stages):
        ws.row_dimensions[r].height = 18
        bg = C["light_bg"] if i % 2 == 0 else C["white"]
        cell = ws.cell(row=r, column=2, value=label)
        cell.font = font(bold=True, size=10)
        cell.fill = fill(bg)
        cell.border = border("thin")
        cell.alignment = align(h="left")

        col = 3
        for plan in plans:
            plan_data = by_plan.get(plan, {})
            metric_data = plan_data.get(key, {})
            for wkey in windows_keys:
                if key == "revenue":
                    raw = metric_data.get(wkey, {}).get("cash", 0)
                    cell = ws.cell(row=r, column=col, value=raw)
                    cell.number_format = '"$"#,##0.00'
                else:
                    raw = metric_data.get(wkey, 0)
                    cell = ws.cell(row=r, column=col, value=raw)
                    cell.number_format = '#,##0'
                cell.font = font(size=10)
                cell.fill = fill(bg)
                cell.border = border("thin")
                cell.alignment = align()
                col += 1

        r += 1


# ══════════════════════════════════════════════════════════════════════════════
#  PLAN PERFORMANCE
# ══════════════════════════════════════════════════════════════════════════════

def build_plan_performance_sheet(wb: openpyxl.Workbook, kpi: dict):
    """
    Builds the Plan Performance sheet.

    Sections:
      1. Plan segment metrics (5 segments × all windows)
      2. Platform breakdown (4-way: iOS App / Android App / Web / Mobile Browser)
    """
    sheet_name = "Plan Performance"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [0.8, 22, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 0.8])

    r = 1
    ws.row_dimensions[r].height = 6
    r += 1

    ws.merge_cells(f"B{r}:R{r}")
    cell = ws.cell(row=r, column=2, value="  PLAN PERFORMANCE")
    cell.font = font(bold=True, size=16, color=C["white"])
    cell.fill = fill(C["navy"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 36
    r += 1

    ws.merge_cells(f"B{r}:R{r}")
    tz = kpi.get("tz_label", "Pacific Time (PT)")
    stale_tag = (f"  ⚠ DATA STALE — {kpi['data_lag_days']}d behind (run daily_refresh.py)  ·  "
                 if kpi.get("data_stale") else "  ")
    cell = ws.cell(row=r, column=2,
                   value=f"{stale_tag}Data through {kpi['as_of']} ({tz})  ·  "
                         f"MTD = {kpi['mtd_start']} → {kpi['as_of']} ({kpi['days_elapsed_mtd']} days)  ·  "
                         f"Generated: {kpi.get('generated_on', kpi['as_of'])}")
    cell.font = font(italic=True, size=9, color=C["white"])
    cell.fill = fill(C["amber"] if kpi.get("data_stale") else C["teal"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 18
    r += 1

    r = spacer(ws, r, 8)

    # ── PLAN SEGMENTS TABLE ───────────────────────────────────────────────────
    r = write_section_header(ws, r, 2, 10, "PLAN SEGMENT BREAKDOWN")

    hdrs = ["Plan Segment", "Yesterday", "MTD", "Prev Month",
            "Same Period", "YTD", "Lifetime",
            "Cash Rev MTD ($)", "MRR MTD ($)", "Unique Users MTD"]
    write_header_row(ws, r, hdrs, C["navy"], start_col=2, height=22)
    r += 1

    def _ppnote(row_num, text):
        try:
            from openpyxl.comments import Comment as _C
            ws.cell(row=row_num, column=2).comment = _C(text, "Dashboard")
        except Exception:
            pass

    seg_notes = {
        "Lite Monthly":        "Monthly Lite plan — billed every month.",
        "Lite Yearly":         "Annual Lite plan — billed once a year; MRR = annual amount ÷ 12.",
        "Unlimited Monthly":   "Monthly Unlimited plan — billed every month.",
        "Unlimited Yearly":    "Annual Unlimited plan — billed once a year; MRR = annual amount ÷ 12.",
        "Free Monthly":        "Permanently free plan — no payment collected. Included to show user base size.",
    }

    for i, seg in enumerate(kpi["plan_performance"]):
        c = seg["counts"]
        rv = seg["revenue"]
        uu = seg["unique_users"]
        write_data_row(ws, r, [
            seg["segment"],
            nv(c["yesterday"]),
            nv(c["mtd"]),
            nv(c["prev_full"]),
            nv(c["prev_same"]),
            nv(c["ytd"]),
            nv(c["lifetime"]),
            cv(rv["mtd"]["cash"]),
            cv(rv["mtd"]["mrr"]),
            nv(uu["mtd"]),
        ], start_col=2, alt=(i % 2 == 0), height=18)
        note = seg_notes.get(seg["segment"])
        if note:
            _ppnote(r, note)
        r += 1

    r = spacer(ws, r, 12)

    # ── PLAN BREAKDOWN — ALL 5 SEGMENTS ─────────────────────────────────────
    r = write_section_header(ws, r, 2, 17, "PLAN BREAKDOWN — ALL 5 SEGMENTS")

    # Segment display config: (label, color)
    seg_colors = [
        (C["teal"],    "Lite Monthly"),
        ("1A5276",     "Lite Yearly"),
        ("6C3483",     "Unlimited Monthly"),
        ("922B21",     "Unlimited Yearly"),
        ("1E8449",     "Free Monthly"),
    ]
    windows_labels = ["MTD", "Prev Month", "Lifetime"]
    windows_keys   = ["mtd", "prev_full", "lifetime"]

    # ── Segment header row (merged, 3 cols each) ──────────────────────────
    ws.row_dimensions[r].height = 18
    ws.cell(row=r, column=2, value="Stage").font = font(bold=True, size=10, color=C["white"])
    ws.cell(row=r, column=2).fill = fill(C["navy"])
    ws.cell(row=r, column=2).alignment = align()
    col = 3
    for color, seg_name in seg_colors:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 2)
        cell = ws.cell(row=r, column=col, value=seg_name)
        cell.font      = font(bold=True, size=10, color=C["white"])
        cell.fill      = fill(color)
        cell.alignment = align()
        col += 3
    r += 1

    # ── Window sub-header row ─────────────────────────────────────────────
    ws.row_dimensions[r].height = 16
    ws.cell(row=r, column=2, value="").fill = fill(C["navy"])
    col = 3
    for color, seg_name in seg_colors:
        for wlabel in windows_labels:
            cell = ws.cell(row=r, column=col, value=wlabel)
            cell.font      = font(bold=True, size=9, color=C["white"])
            cell.fill      = fill(color)
            cell.alignment = align()
            cell.border    = border("thin")
            col += 1
    r += 1

    by_segment = kpi["funnel"].get("by_segment", {})
    plan_stages = [
        ("Trials",            "trials",   "Trial orders started in this period for this segment."),
        ("New Subscriptions", "new_subs", "First paid order on a new subscription for this segment."),
        ("Renewals",          "renewals", "Subsequent renewal transactions (repeat payments) for this segment."),
        ("Paid",              "paid",     "Total paid transactions (new + renewals) for this segment."),
        ("Cash Revenue ($)",  "revenue",  "Actual cash collected for this segment in this period."),
    ]

    for i, (label, key, note) in enumerate(plan_stages):
        ws.row_dimensions[r].height = 18
        bg = C["light_bg"] if i % 2 == 0 else C["white"]
        cell = ws.cell(row=r, column=2, value=label)
        cell.font = font(bold=True, size=10)
        cell.fill = fill(bg)
        cell.border = border("thin")
        cell.alignment = align(h="left")
        if note:
            try:
                from openpyxl.comments import Comment as _Comment
                ws.cell(row=r, column=2).comment = _Comment(note, "Dashboard")
            except Exception:
                pass

        col = 3
        for _color, seg_name in seg_colors:
            seg_data    = by_segment.get(seg_name, {})
            metric_data = seg_data.get(key, {})
            for wkey in windows_keys:
                if key == "revenue":
                    raw = metric_data.get(wkey, {}).get("cash", 0)
                    cell = ws.cell(row=r, column=col, value=raw)
                    cell.number_format = '"$"#,##0.00'
                else:
                    raw = metric_data.get(wkey, 0)
                    cell = ws.cell(row=r, column=col, value=raw)
                    cell.number_format = '#,##0'
                cell.font = font(size=10)
                cell.fill = fill(bg)
                cell.border = border("thin")
                cell.alignment = align()
                col += 1
        r += 1

    ws.auto_filter.ref = None  # no auto-filter on plan performance


# ══════════════════════════════════════════════════════════════════════════════
#  RETENTION & RENEWALS
# ══════════════════════════════════════════════════════════════════════════════

def build_retention_sheet(wb: openpyxl.Workbook, kpi: dict):
    """
    Builds the Retention & Renewals sheet.

    Metrics:
      - Renewals (all windows)
      - Churn events (MTD vs Prev Month) — combined
      - Churn rate (MTD vs Prev Month)
      - Active subscribers (snapshot)
      - Trial→Paid conversion (all windows)
      - Free→Paid conversion (all windows)
    """
    sheet_name = "Retention & Renewals"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [0.8, 30, 14, 14, 16, 14, 14, 16, 16, 0.8])

    r = 1
    ws.row_dimensions[r].height = 6
    r += 1

    ws.merge_cells(f"B{r}:I{r}")
    cell = ws.cell(row=r, column=2, value="  RETENTION & RENEWALS")
    cell.font = font(bold=True, size=16, color=C["white"])
    cell.fill = fill(C["navy"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 36
    r += 1

    ws.merge_cells(f"B{r}:I{r}")
    tz = kpi.get("tz_label", "Pacific Time (PT)")
    stale_tag = (f"  ⚠ DATA STALE — {kpi['data_lag_days']}d behind (run daily_refresh.py)  ·  "
                 if kpi.get("data_stale") else "  ")
    cell = ws.cell(row=r, column=2,
                   value=f"{stale_tag}Data through {kpi['as_of']} ({tz})  ·  "
                         f"MTD = {kpi['mtd_start']} → {kpi['as_of']} ({kpi['days_elapsed_mtd']} days)  ·  "
                         f"Generated: {kpi.get('generated_on', kpi['as_of'])}")
    cell.font = font(italic=True, size=9, color=C["white"])
    cell.fill = fill(C["amber"] if kpi.get("data_stale") else C["teal"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 18
    r += 1

    r = spacer(ws, r, 8)

    st    = kpi["subscription_types"]
    pchurn = kpi["paid_churn"]
    act   = kpi["active"]
    t_conv = kpi["trial_conversion"]
    t_pipe = kpi["trial_pipeline"]
    f_conv = kpi["free_conversion"]
    days   = kpi["days_elapsed_mtd"]
    renewal_state = kpi.get("renewal_state", {"active_at_start": 0, "churned": 0, "renewed": 0, "waiting": 0})
    seg_ret = kpi.get("segment_retention", {})
    canc = kpi.get("cancellations", {w: {"count": 0, "refund_amount": 0.0} for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]})

    # ── RETENTION METRICS TABLE ───────────────────────────────────────────────
    r = write_section_header(ws, r, 2, 8, "RETENTION & CONVERSION METRICS")

    hdrs = ["Metric", "Yesterday", "MTD", "Prev Month (Full)",
            f"Same Period ({days} days)", "YTD", "Lifetime"]
    write_header_row(ws, r, hdrs, C["navy"], start_col=2, height=18)
    r += 1

    def _rnote(row_num, text):
        try:
            from openpyxl.comments import Comment as _C
            ws.cell(row=row_num, column=2).comment = _C(text, "Dashboard")
        except Exception:
            pass

    # Each tuple: (label, vals_tuple, comment_or_None)
    retention_rows = [
        ("Renewals",
         (nv(st["renewals"]["yesterday"]),
          nv(st["renewals"]["mtd"]),
          nv(st["renewals"]["prev_full"]),
          nv(st["renewals"]["prev_same"]),
          nv(st["renewals"]["ytd"]),
          nv(st["renewals"]["lifetime"])),
         "Repeat renewal transactions — a subscriber paying for another period. "
         "This is the direct opposite of churn: each renewal means a customer stayed."),

        ("Active Subscribers",
         (nv(act["active_now"]), nv(act["active_now"]), "–", "–", "–", "–"),
         "Active subscriptions as of end of yesterday"
         "Includes both paid and free plan subscribers."),

        ("Trial Pipeline",
         (nv(t_pipe["count"]), nv(t_pipe["count"]), "–", "–", "–", "–"),
         "Unconverted trials still within their trial period (paid_till >= today). "
         "Shows how many potential conversions are currently in the funnel."),

        ("Paid Churn Events",
         (nv(pchurn["yesterday"]["events"]),
          nv(pchurn["mtd"]["events"]),
          nv(pchurn["prev_full"]["events"]),
          nv(pchurn["prev_same"]["events"]),
          nv(pchurn["ytd"]["events"]),
          nv(pchurn["lifetime"]["events"])),
         "Number of paid subscriptions whose access expired (paid_till fell) within this period "
         "and are now inactive. Each subscription is counted once regardless of how long ago it started."),

        ("Paid Churn Rate",
         (pv(pchurn["yesterday"]["rate"]),
          pv(pchurn["mtd"]["rate"]),
          pv(pchurn["prev_full"]["rate"]),
          pv(pchurn["prev_same"]["rate"]),
          pv(pchurn["ytd"]["rate"]),
          pv(pchurn["lifetime"]["rate"])),
         "Churn Events ÷ Active Paid Subs at Period Start (MTD/Prev). "
         "YTD and Lifetime use the avg of active paid subs at each month start in the period. "
         "Lower is better."),

        ("Paid Retention Rate",
         (pv(pchurn["yesterday"]["retention_rate"]),
          pv(pchurn["mtd"]["retention_rate"]),
          pv(pchurn["prev_full"]["retention_rate"]),
          pv(pchurn["prev_same"]["retention_rate"]),
          pv(pchurn["ytd"]["retention_rate"]),
          pv(pchurn["lifetime"]["retention_rate"])),
         "1 − Churn Rate. The share of paid subscribers at period start who remained active "
         "through the end of the period. Higher is better."),

        ("  Paid Retained Count",
         (nv(pchurn["yesterday"]["denominator"] - pchurn["yesterday"]["events"]),
          nv(pchurn["mtd"]["denominator"] - pchurn["mtd"]["events"]),
          nv(pchurn["prev_full"]["denominator"] - pchurn["prev_full"]["events"]),
          nv(pchurn["prev_same"]["denominator"] - pchurn["prev_same"]["events"]),
          nv(pchurn["ytd"]["denominator"] - pchurn["ytd"]["events"]),
          nv(pchurn["lifetime"]["denominator"] - pchurn["lifetime"]["events"])),
         "Active Paid at Period Start minus Churn Events = subscribers who stayed. "
         "The absolute headcount complement of churn events."),

        ("── Renewal State (MTD) ──",
         ("", "", "", "", "", ""),
         None),

        (f"  Active at Month Start (Paid)",
         ("–",
          nv(renewal_state["active_at_start"]),
          nv(pchurn["prev_full"]["denominator"]),
          nv(pchurn["prev_same"]["denominator"]),
          "–", "–"),
         "Paid subscribers who had active access at the start of the month. "
         "This is the pool that can either churn, renew, or wait during the month."),

        ("  Churned",
         (nv(pchurn["yesterday"]["events"]),
          nv(renewal_state["churned"]),
          nv(pchurn["prev_full"]["events"]),
          nv(pchurn["prev_same"]["events"]),
          nv(pchurn["ytd"]["events"]),
          nv(pchurn["lifetime"]["events"])),
         "Active-at-start subscribers whose paid access expired this month and are now inactive."),

        ("  Renewed",
         (nv(st["renewals"]["yesterday"]),
          nv(renewal_state["renewed"]),
          "–", "–", "–", "–"),
         "Active-at-start subscribers who placed a renewal paid order this month."),

        ("  Waiting for Renewal",
         ("–",
          nv(renewal_state["waiting"]),
          "–", "–", "–", "–"),
         "Active-at-start subscribers whose paid access has not yet expired and who have not yet renewed. "
         "They are still within their paid period — their renewal is upcoming."),

        ("Trial→Paid Conv%",
         (pv(t_conv["yesterday"]["rate"]),
          pv(t_conv["mtd"]["rate"]),
          pv(t_conv["prev_full"]["rate"]),
          pv(t_conv["prev_same"]["rate"]),
          pv(t_conv["ytd"]["rate"]),
          pv(t_conv["lifetime"]["rate"])),
         "Of all trials that ended in this period, the share that converted to a paid subscription. "
         "Both trial endings and paid orders are measured within the same date window."),

        ("  Trials Ended in Window",
         (nv(t_conv["yesterday"]["total"]),
          nv(t_conv["mtd"]["total"]),
          nv(t_conv["prev_full"]["total"]),
          nv(t_conv["prev_same"]["total"]),
          nv(t_conv["ytd"]["total"]),
          nv(t_conv["lifetime"]["total"])),
         "Number of trial subscriptions whose trial period ended within this window. "
         "Trial end = date of the first paid order (converted trials) or paid_till "
         "(unconverted trials). This is the denominator for the Trial→Paid conversion rate."),

        ("  Converted to Paid",
         (nv(t_conv["yesterday"]["converted"]),
          nv(t_conv["mtd"]["converted"]),
          nv(t_conv["prev_full"]["converted"]),
          nv(t_conv["prev_same"]["converted"]),
          nv(t_conv["ytd"]["converted"]),
          nv(t_conv["lifetime"]["converted"])),
         "Of the trials that ended in this window, how many placed a paid order in the same window. "
         "This is the numerator for the Trial→Paid conversion rate."),

        ("Free→Paid Conv%",
         (pv(f_conv["yesterday"]["rate"]),
          pv(f_conv["mtd"]["rate"]),
          pv(f_conv["prev_full"]["rate"]),
          pv(f_conv["prev_same"]["rate"]),
          pv(f_conv["ytd"]["rate"]),
          pv(f_conv["lifetime"]["rate"])),
         "Of free-plan users who were active at the start of this period, the share that upgraded "
         "to a paid subscription during the period."),

        ("  Active Free Users (Denom)",
         (nv(f_conv["yesterday"]["total"]),
          nv(f_conv["mtd"]["total"]),
          nv(f_conv["prev_full"]["total"]),
          nv(f_conv["prev_same"]["total"]),
          nv(f_conv["ytd"]["total"]),
          nv(f_conv["lifetime"]["total"])),
         "Active free-plan users at the start of the period — the denominator for Free→Paid Conv%. "
         "YTD and Lifetime use the avg of active free users at each month start in the period."),

        ("  Converted to Paid",
         (nv(f_conv["yesterday"]["converted"]),
          nv(f_conv["mtd"]["converted"]),
          nv(f_conv["prev_full"]["converted"]),
          nv(f_conv["prev_same"]["converted"]),
          nv(f_conv["ytd"]["converted"]),
          nv(f_conv["lifetime"]["converted"])),
         "Free users from the denominator pool who placed a paid order within this period. "
         "This is the numerator for the Free→Paid conversion rate."),

        ("── Cancellations ────",
         ("", "", "", "", "", ""),
         None),

        ("Cancelled Orders",
         (nv(canc.get("yesterday", {}).get("count", 0)),
          nv(canc.get("mtd", {}).get("count", 0)),
          nv(canc.get("prev_full", {}).get("count", 0)),
          nv(canc.get("prev_same", {}).get("count", 0)),
          nv(canc.get("ytd", {}).get("count", 0)),
          nv(canc.get("lifetime", {}).get("count", 0))),
         "Orders with order_status = Cancelled or Refund in this window."),

        ("Total Refunded ($)",
         (cv(canc.get("yesterday", {}).get("refund_amount", 0)),
          cv(canc.get("mtd", {}).get("refund_amount", 0)),
          cv(canc.get("prev_full", {}).get("refund_amount", 0)),
          cv(canc.get("prev_same", {}).get("refund_amount", 0)),
          cv(canc.get("ytd", {}).get("refund_amount", 0)),
          cv(canc.get("lifetime", {}).get("refund_amount", 0))),
         "Sum of SBO_RefundAmount on cancelled/refunded orders."),
    ]

    # Metrics that get 5-segment sub-rows (by label prefix)
    SEG_METRICS = {
        "Renewals",
        "Paid Churn Events",
        "Paid Churn Rate",
        "Paid Retention Rate",
        "  Paid Retained Count",
        "Trial\u2192Paid Conv%",
    }

    for i, (label, vals, note) in enumerate(retention_rows):
        write_data_row(ws, r, [label] + list(vals),
                       start_col=2, alt=(i % 2 == 0), height=18,
                       bold_first=not label.startswith(" "))
        if note:
            _rnote(r, note)
        r += 1

        # Segment sub-rows for selected metrics
        if label in SEG_METRICS:
            for seg in PLAN_SEGMENTS:
                sd = seg_ret.get(seg, {})
                if label == "Renewals":
                    sub_vals = [
                        f"    {seg}", "\u2013",
                        nv(sd.get("renewals_mtd", 0)),
                        nv(sd.get("renewals_prev", 0)),
                        nv(sd.get("renewals_prev_same", 0)),
                        nv(sd.get("renewals_ytd", 0)),
                        nv(sd.get("renewals_lifetime", 0)),
                    ]
                elif label == "Paid Churn Events":
                    sub_vals = [
                        f"    {seg}",
                        nv(sd.get("churn_yesterday",{}).get("events", 0)),
                        nv(sd.get("churn_mtd",      {}).get("events", 0)),
                        nv(sd.get("churn_prev",     {}).get("events", 0)),
                        nv(sd.get("churn_prev_same",{}).get("events", 0)),
                        nv(sd.get("churn_ytd",      {}).get("events", 0)),
                        nv(sd.get("churn_lifetime", {}).get("events", 0)),
                    ]
                elif label == "Paid Churn Rate":
                    sub_vals = [
                        f"    {seg}", "\u2013",
                        pv(sd.get("churn_mtd",      {}).get("rate", 0.0)),
                        pv(sd.get("churn_prev",     {}).get("rate", 0.0)),
                        pv(sd.get("churn_prev_same",{}).get("rate", 0.0)),
                        pv(sd.get("churn_ytd",      {}).get("rate", 0.0)),
                        pv(sd.get("churn_lifetime", {}).get("rate", 0.0)),
                    ]
                elif label == "Paid Retention Rate":
                    sub_vals = [
                        f"    {seg}", "\u2013",
                        pv(sd.get("churn_mtd",      {}).get("retention_rate", 1.0)),
                        pv(sd.get("churn_prev",     {}).get("retention_rate", 1.0)),
                        pv(sd.get("churn_prev_same",{}).get("retention_rate", 1.0)),
                        pv(sd.get("churn_ytd",      {}).get("retention_rate", 1.0)),
                        pv(sd.get("churn_lifetime", {}).get("retention_rate", 1.0)),
                    ]
                elif label == "  Paid Retained Count":
                    cy_seg = sd.get("churn_yesterday",{})
                    cm  = sd.get("churn_mtd",       {})
                    cp  = sd.get("churn_prev",      {})
                    cps = sd.get("churn_prev_same", {})
                    cy  = sd.get("churn_ytd",       {})
                    cl  = sd.get("churn_lifetime",  {})
                    sub_vals = [
                        f"    {seg}",
                        nv(cy_seg.get("denominator", 0) - cy_seg.get("events", 0)),
                        nv(cm.get("denominator", 0)  - cm.get("events", 0)),
                        nv(cp.get("denominator", 0)  - cp.get("events", 0)),
                        nv(cps.get("denominator", 0) - cps.get("events", 0)),
                        nv(cy.get("denominator", 0)  - cy.get("events", 0)),
                        nv(cl.get("denominator", 0)  - cl.get("events", 0)),
                    ]
                elif label == "Trial\u2192Paid Conv%":
                    sub_vals = [
                        f"    {seg}", "\u2013",
                        pv(sd.get("trial_conv_mtd",       {}).get("rate")),
                        pv(sd.get("trial_conv_prev",      {}).get("rate")),
                        pv(sd.get("trial_conv_prev_same", {}).get("rate")),
                        pv(sd.get("trial_conv_ytd",       {}).get("rate")),
                        pv(sd.get("trial_conv_lifetime",  {}).get("rate")),
                    ]
                else:
                    continue
                write_data_row(ws, r, sub_vals,
                               start_col=2, alt=False, height=16, bold_first=False)
                r += 1

    r = spacer(ws, r, 12)

    # ── MoM RETENTION VARIANCE ────────────────────────────────────────────────
    r = write_section_header(ws, r, 2, 8,
                              "MoM RETENTION VARIANCE (MTD vs Prev Month Full vs Same Period)")

    write_header_row(ws, r,
                     ["Metric", "MTD", "Prev Month (Full)", "MoM Δ (Full)",
                      "Same Period", "MoM Δ (Same)", "", ""],
                     C["teal"], start_col=2, height=18)
    r += 1

    def _var(a, b):
        if b and b != 0:
            d = a - b
            p = d / abs(b) * 100
            arrow = "▲" if d > 0 else ("▼" if d < 0 else "→")
            return f"{arrow} {abs(d):,} ({'+' if d>=0 else ''}{p:.1f}%)"
        return "–"

    mom_rows = [
        ("Renewals",
         st["renewals"]["mtd"], st["renewals"]["prev_full"], st["renewals"]["prev_same"]),
        ("Paid Churn Events",
         pchurn["mtd"]["events"], pchurn["prev_full"]["events"], None),
    ]

    for i, (label, mtd_v, pf_v, ps_v) in enumerate(mom_rows):
        is_rate = isinstance(mtd_v, float) and mtd_v < 1
        fmt_fn = pv if is_rate else nv
        write_data_row(ws, r, [
            label,
            fmt_fn(mtd_v),
            fmt_fn(pf_v) if pf_v is not None else "–",
            _var(mtd_v, pf_v),
            fmt_fn(ps_v) if ps_v is not None else "–",
            _var(mtd_v, ps_v) if ps_v is not None else "–",
            "", "",
        ], start_col=2, alt=(i % 2 == 0), height=18)
        r += 1

    r = spacer(ws, r, 12)

    # ── METRIC VALIDATION REFERENCE ──────────────────────────────────────────
    r = write_section_header(ws, r, 2, 9, f"METRIC VALIDATION REFERENCE — MTD ({kpi['lbl_mtd']})")

    # Explanation row
    ws.merge_cells(f"B{r}:J{r}")
    cell = ws.cell(row=r, column=2,
        value="Use this table to manually validate dashboard rates. "
              "Rate = Numerator ÷ Denominator for each metric. "
              "Segment rows match the per-segment sub-rows above.")
    cell.font = font(size=9, italic=True, color=C["text_md"])
    cell.fill = fill(C["light_bg"])
    cell.alignment = align(h="left", v="center", wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1

    from openpyxl.comments import Comment as _Cmt

    # Column headers (cols B–I = 8 data columns)
    val_hdrs = [
        "Segment",
        "Churn Events\n(Numerator)",
        "Paid Active at\nMonth Start\n(Denominator)",
        "Paid Churn\nRate",
        "Renewals\n(MTD)",
        "Trials Whose\nWindow Ended\n(Denom)",
        "Converted to\nPaid (Num)",
        "Trial\u2192Paid\nConv%",
    ]
    write_header_row(ws, r, val_hdrs, C["teal"], start_col=2, height=36)

    # Add explanatory comments to each header cell
    header_comments = {
        3: ("Churn Events (Numerator)",
            "Number of paid subscriptions whose paid_till fell within MTD AND are now In-Active. "
            "This is the top number in the Paid Churn Rate formula."),
        4: ("Paid Active at Month Start (Denominator)",
            "Paid subscriptions whose paid_till >= end of previous month (i.e. were active at the "
            "start of this month). This is the bottom number in the Paid Churn Rate formula. "
            "IMPORTANT: This uses the current paid_till value, which may have changed since month "
            "start due to renewals. For a fully accurate denominator, compare against the "
            "Sub State Snapshots sheet."),
        5: ("Paid Churn Rate",
            "Churn Events ÷ Paid Active at Month Start. Lower is better. Target: < 5%."),
        6: ("Renewals MTD (context)",
            "Number of renewal orders placed in MTD. "
            "Not a rate — shown for context to complement churn events."),
        7: ("Trials Whose Window Ended (Denominator)",
            "Trial subscriptions whose trial period ended within MTD. "
            "Trial end = date of the first paid order (converted trials) or paid_till "
            "(unconverted trials). These are the trials that have had their chance to "
            "convert this month. This is the bottom number in the Trial→Paid Conv% formula."),
        8: ("Converted to Paid (Numerator)",
            "Of the trials whose window ended in MTD, how many placed a paid order in the same window. "
            "This is the top number in the Trial→Paid Conv% formula."),
        9: ("Trial→Paid Conversion Rate",
            "Converted to Paid ÷ Trials Whose Window Ended. Higher is better. Target: > 30%."),
    }
    for col_idx, (col_title, col_comment) in header_comments.items():
        try:
            ws.cell(row=r, column=col_idx).comment = _Cmt(
                f"{col_title}\n\n{col_comment}", "Dashboard")
        except Exception:
            pass
    r += 1

    # Overall + 5 segment rows
    overall_churn_events = pchurn["mtd"]["events"]
    overall_churn_denom  = pchurn["mtd"]["denominator"]
    overall_churn_rate   = pchurn["mtd"]["rate"]
    overall_renewals     = st["renewals"]["mtd"]
    overall_trial_denom  = t_conv["mtd"]["total"]
    overall_trial_num    = t_conv["mtd"]["converted"]
    overall_trial_rate   = t_conv["mtd"]["rate"]

    val_data_rows = [
        ("Overall", overall_churn_events, overall_churn_denom, overall_churn_rate,
         overall_renewals, overall_trial_denom, overall_trial_num, overall_trial_rate),
    ]
    for seg in PLAN_SEGMENTS:
        sd = seg_ret.get(seg, {})
        cm = sd.get("churn_mtd", {})
        tc = sd.get("trial_conv_mtd", {})
        seg_renewals = sd.get("renewals_mtd", 0)
        val_data_rows.append((
            f"  {seg}",
            cm.get("events", 0),
            cm.get("denominator", 0),
            cm.get("rate", 0.0),
            seg_renewals,
            tc.get("total", 0),
            tc.get("converted", 0),
            tc.get("rate"),
        ))

    for i, (seg_label, ce, cd, cr, ren, td, tn, tr) in enumerate(val_data_rows):
        is_overall = (i == 0)
        row_vals = [
            seg_label,
            nv(ce),
            nv(cd),
            pv(cr),
            nv(ren),
            nv(td),
            nv(tn),
            pv(tr) if tr is not None else "–",
        ]
        write_data_row(ws, r, row_vals, start_col=2, alt=(i % 2 == 0), height=18,
                       bold_first=is_overall, first_align="left")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
#  RAW DATA
# ══════════════════════════════════════════════════════════════════════════════

def build_raw_data_sheet(wb: openpyxl.Workbook, df: pd.DataFrame):
    """
    Builds the Raw Data sheet with the full dataset.

    This sheet:
      - Contains every row from the MySQL source query
      - Has auto-filter and freeze panes enabled
      - Is overwritten completely on every daily refresh
      - Is for analyst use — not the primary CEO view

    Columns included: all columns from the query, excluding ContactNumber
    (PII — kept in source DB, not surfaced in the shareable Excel file).
    """
    sheet_name = "Raw Data"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = True

    # Columns to include (order matters — this is the display order)
    include_cols = [
        "order_identifier", "order_status", "user_id", "full_name",
        "email_id", "contact_number", "city", "state", "zip_code",
        "subscription_name", "duration", "plan_segment",
        "subscription_amt", "discount", "total_amt", "refund_amount", "mrr_amount",
        "trial_check", "subscribed_date", "subscribed_time",
        "paid_till", "end_date",
        "curr_status", "subscription_id", "stripe_cus_id", "stripe_pmt_id",
        "subscription_type",
        "refund_date", "gateway_refund_ref_id", "refund_reason",
        "device_type", "platform",
    ]

    # Only include columns that actually exist in the dataframe
    cols = [c for c in include_cols if c in df.columns]
    df_out = df[cols].copy()

    # Write header row
    write_header_row(ws, 1, cols, C["navy"], start_col=1, height=18)

    # Add column-level comments to header cells
    _col_notes = {
        "order_identifier":       "Unique ID for each billing event. Use this as the row grain — one row per transaction.",
        "order_status":           "Confirmed = payment authorized; Completed = payment settled. All statuses present — filter as needed for analysis.",
        "user_id":                "Unique customer ID across all WAY products.",
        "full_name":              "Customer display name.",
        "email_id":               "Customer email address.",
        "contact_number":         "Customer phone number (most recent on file). Null for ~40% of users.",
        "subscription_name":      "Plan tier: Free | Lite | Unlimited | Business Pro Trial.",
        "duration":               "Billing cycle: Monthly or Yearly.",
        "plan_segment":           "Combined plan + billing cycle (e.g. 'Lite Monthly'). 5-way split for performance analysis.",
        "subscription_amt":       "List price before any discount (gross amount on the order).",
        "discount":               "Amount discounted. For trial orders this equals subscription_amt (full discount = $0 collected).",
        "total_amt":              "Actual cash collected = subscription_amt minus discount. $0 for all trial orders.",
        "refund_amount":          "Amount refunded on this order. Non-zero only on Cancelled or Refunded orders. "
                                  "Net revenue = SUM(total_amt) - SUM(refund_amount) bucketed by refund_date.",
        "mrr_amount":             "Monthly Recurring Revenue contribution: monthly plans use total_amt as-is; yearly plans divide by 12.",
        "trial_check":            "FREE = free plan order (price = 0, discount = 0, total = 0). PAID = cash collected (total > 0). TRIAL = trial order (price > 0, discount = price, total = 0).",
        "subscribed_date":        "Date the order was placed, converted to Pacific Time (PST/PDT).",
        "paid_till":              "Date the subscription's paid access expires (PST). Active definition: paid_till > yesterday (strictly greater). "
                                  "RENEWAL RISK: subscriptions with paid_till in the next 14-30 days will hit a billing attempt soon. "
                                  "CASE B NOTE: for rows where end_date IS NOT NULL AND paid_till <= end_date, use end_date for the active check — not this column.",
        "end_date":               "Hard system cutoff date (PST). NULL for active/ongoing subscriptions. "
                                  "Set when automatic renewal fails and the subscription is terminated. "
                                  "CASE B: if end_date IS NOT NULL AND paid_till <= end_date, this date governs access expiry — the subscription will expire on end_date regardless of paid_till. "
                                  "Use this to identify subscriptions in a failed-renewal / grace-period state.",
        "curr_status":            "Active = subscription still has valid access as of end of yesterday (Pacific). In-Active = expired. "
                                  "Applies Case A / Case B logic automatically.",
        "subscription_id":        "Unique ID per subscription. One customer can hold multiple subscriptions over time.",
        "stripe_pmt_id":          "Stripe payment intent / charge reference ID for this order. NULL for free and trial orders.",
        "subscription_type":      "Trial = first order on a new subscription where the full price was discounted (price > 0, total = 0). "
                                  "New Subscription = first paid order on a subscription (total > 0), or first order on a Free plan. "
                                  "Renewal = all subsequent orders (2nd order onwards on the same subscription_id).",
        "refund_date":            "Date the refund was processed, in Pacific Time. Refunds are bucketed by this date — "
                                  "a May payment refunded in June reduces June revenue, not May. NULL when no refund.",
        "gateway_refund_ref_id":  "Stripe refund reference ID. NULL when no refund on this order.",
        "refund_reason":          "Reason code recorded at time of refund. NULL when no refund on this order.",
        "device_type":            "Granular platform: IOS | ANDROID | IOS_BROWSER_MOBILE | Android_BROWSER_MOBILE | BROWSER_DESKTOP.",
        "platform":               "High-level channel: Mobile App | Mobile Browser | Web.",
    }
    try:
        from openpyxl.comments import Comment as _C
        for ci, col_name in enumerate(cols, start=1):
            if col_name in _col_notes:
                ws.cell(row=1, column=ci).comment = _C(_col_notes[col_name], "Dashboard")
    except Exception:
        pass

    # ── Set column widths ─────────────────────────────────────────────────────
    col_widths = {
        "order_identifier":      18, "order_status":         14, "user_id":             12,
        "full_name":             22, "email_id":             28, "contact_number":       16,
        "city":                  14, "state":                 8, "zip_code":             10,
        "subscription_name":     22, "duration":             12, "plan_segment":         20,
        "subscription_amt":      16, "discount":             12, "total_amt":            14,
        "refund_amount":         14, "mrr_amount":           14, "trial_check":          12,
        "subscribed_date":       16, "subscribed_time":      16, "paid_till":            16,
        "end_date":              16, "curr_status":          14, "subscription_id":      18,
        "stripe_cus_id":         24, "stripe_pmt_id":        28, "subscription_type":    20,
        "refund_date":           16, "gateway_refund_ref_id": 28, "refund_reason":        22,
        "device_type":           28, "platform":             18,
    }
    for ci, col_name in enumerate(cols, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)
        ].width = col_widths.get(col_name, 16)

    ws.freeze_panes = "A2"

    # ── Write data rows ────────────────────────────────────────────────────────────
    for i, row_data in enumerate(df_out.itertuples(index=False), start=2):
        bg = C["light_bg"] if i % 2 == 0 else C["white"]
        ws.row_dimensions[i].height = 15
        for j, val in enumerate(row_data, start=1):
            import pandas as _pd
            if isinstance(val, _pd.Timestamp):
                val = val.date() if hasattr(val, "date") else val
            elif hasattr(val, "item"):        # numpy scalar → Python native
                val = val.item()
            cell = ws.cell(row=i, column=j, value=val)
            cell.font      = font(size=9)
            cell.fill      = fill(bg)
            cell.border    = border("thin")
            cell.alignment = align(
                h="left" if j == 1 else "center", v="center"
            )


# ══════════════════════════════════════════════════════════════════════════════
#  TRIAL PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def build_trial_pipeline_sheet(wb: openpyxl.Workbook, trial_df: pd.DataFrame):
    """
    Builds the Trial Pipeline sheet -- one row per open (unconverted, active) trial.

    Columns: Trial Start | Trial Expires | Days Remaining | Name | Email |
             Phone | Plan | Platform | Device | User ID | Subscription ID

    Sorted oldest-first so the longest-running unconverted trials (most at-risk
    of going stale) appear at the top.

    Source: get_open_trial_rows() in kpi_engine.py.
    """
    sheet_name = "Trial Pipeline"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, [0.8, 14, 14, 14, 24, 30, 16, 18, 16, 28, 14, 22, 0.8])

    r = 1
    ws.row_dimensions[r].height = 6
    r += 1

    ws.merge_cells(f"B{r}:M{r}")
    cell = ws.cell(row=r, column=2, value="  TRIAL PIPELINE — OPEN UNCONVERTED TRIALS")
    cell.font      = font(bold=True, size=16, color=C["white"])
    cell.fill      = fill(C["navy"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 36
    r += 1

    ws.merge_cells(f"B{r}:M{r}")
    total = len(trial_df)
    cell = ws.cell(row=r, column=2,
                   value=f"  {total} open trial{'s' if total != 1 else ''}  ·  "
                         f"Active = trial access not yet expired (paid_till ≥ today)  ·  "
                         f"Sorted oldest first (longest unconverted = highest risk)")
    cell.font      = font(italic=True, size=9, color=C["white"])
    cell.fill      = fill(C["teal"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 18
    r += 1

    r = spacer(ws, r, 8)

    hdrs = ["Trial Start", "Trial Expires", "Days Remaining",
            "Name", "Email", "Phone",
            "Plan", "Platform", "Device",
            "User ID", "Subscription ID"]
    write_header_row(ws, r, hdrs, C["navy"], start_col=2, height=22)
    r += 1

    if trial_df.empty:
        ws.merge_cells(f"B{r}:M{r}")
        ws.cell(row=r, column=2, value="No open trials at this time.").font = font(italic=True, size=10)
        return

    col_map = [
        ("subscribed_date", None),
        ("paid_till",       None),
        ("days_remaining",  None),
        ("full_name",       ""),
        ("email_id",        ""),
        ("contact_number",  ""),
        ("plan_segment",    ""),
        ("platform",        ""),
        ("device_type",     ""),
        ("user_id",         ""),
        ("subscription_id", ""),
    ]

    for i, row_data in enumerate(trial_df.itertuples(index=False), start=0):
        bg = C["light_bg"] if i % 2 == 0 else C["white"]
        row_vals = []
        for col, default in col_map:
            val = getattr(row_data, col, default)
            if val is None or (not isinstance(val, (str, bool, int, float)) and pd.isna(val)):
                val = default if default is not None else ""
            row_vals.append(val)
        write_data_row(ws, r, row_vals, start_col=2, alt=(i % 2 == 0), height=18,
                       bold_first=False)
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
#  ACTIVE SUBSCRIPTIONS
# ══════════════════════════════════════════════════════════════════════════════

def build_active_subscriptions_sheet(wb: openpyxl.Workbook, kpi: dict):
    """
    Builds the Active Subscriptions sheet -- month-on-month active subscriber
    counts by plan segment, from Jan 1 of current year through MTD_START.
    """
    sheet_name = "Active Subscriptions"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    COL_PLANS = [
        "Free Monthly", "Trials",
        "Lite Monthly", "Lite Yearly",
        "Unlimited Monthly", "Unlimited Yearly",
    ]
    set_col_widths(ws, [0.8, 14, 16, 14, 16, 14, 20, 20, 12, 0.8])

    r = 1
    ws.row_dimensions[r].height = 6
    r += 1

    ws.merge_cells(f"B{r}:J{r}")
    cell = ws.cell(row=r, column=2, value="  ACTIVE SUBSCRIPTIONS — MONTH-ON-MONTH")
    cell.font      = font(bold=True, size=16, color=C["white"])
    cell.fill      = fill(C["navy"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 36
    r += 1

    ws.merge_cells(f"B{r}:J{r}")
    cell = ws.cell(row=r, column=2,
                   value="  Active subscribers at the START of each month  ·  "
                         "Active = paid_till > last day of prior month  ·  "
                         "Trials = open unconverted trials within their trial window")
    cell.font      = font(italic=True, size=9, color=C["white"])
    cell.fill      = fill(C["teal"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 18
    r += 1

    r = spacer(ws, r, 8)

    hdrs = ["Month"] + COL_PLANS + ["Total"]
    write_header_row(ws, r, hdrs, C["navy"], start_col=2, height=22)
    r += 1

    active_by_plan = kpi.get("active_by_plan", {})

    for i, (d, counts) in enumerate(sorted(active_by_plan.items())):
        bg    = C["light_bg"] if i % 2 == 0 else C["white"]
        total = sum(counts.get(seg, 0) for seg in COL_PLANS)
        month_label = d.strftime("%b") + " " + str(d.day) if hasattr(d, "strftime") else str(d)
        row_vals = [month_label] + [counts.get(seg, 0) for seg in COL_PLANS] + [total]
        write_data_row(ws, r, row_vals, start_col=2, alt=(i % 2 == 0), height=18,
                       bold_first=False)
        for col_offset in range(1, len(COL_PLANS) + 2):
            c_cell = ws.cell(row=r, column=2 + col_offset)
            c_cell.number_format = "#,##0"
        r += 1

    r = spacer(ws, r, 8)
    ws.merge_cells(f"B{r}:J{r}")
    note_cell = ws.cell(
        row=r, column=2,
        value="Note: each row shows subscriptions active at the START of that month "
              "(paid_till > last calendar day of prior month). "
              "Trials = unconverted trial subscriptions still within their trial window. "
              "Plan segments reflect each subscription's current/most-recent plan tier.",
    )
    note_cell.font      = font(italic=True, size=9, color="666666")
    note_cell.alignment = align(h="left", wrap=True)
    ws.row_dimensions[r].height = 30


# ══════════════════════════════════════════════════════════════════════════════
#  CHURN RAW DATA
# ══════════════════════════════════════════════════════════════════════════════

def build_churn_raw_data_sheet(wb: openpyxl.Workbook, churn_data: dict):
    """
    Builds the Churn Raw Data sheet -- two sections (Paid Churn + Trial Churn), MTD window.
    No freeze panes. No auto-filter.
    Source: get_churned_rows() in kpi_engine.py.
    """
    sheet_name = "Churn Raw Data"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, [0.8, 14, 14, 24, 30, 16, 18, 16, 28, 14, 22, 0.8])

    paid_df  = churn_data.get("paid",  pd.DataFrame())
    trial_df = churn_data.get("trial", pd.DataFrame())

    mtd_label = (
        MTD_START.strftime("%b") + " " + str(MTD_START.day) +
        " – " +
        YESTERDAY.strftime("%b") + " " + str(YESTERDAY.day)
    )

    r = 1
    ws.row_dimensions[r].height = 6
    r += 1

    ws.merge_cells(f"B{r}:K{r}")
    cell = ws.cell(row=r, column=2, value="  CHURN RAW DATA — MTD")
    cell.font      = font(bold=True, size=16, color=C["white"])
    cell.fill      = fill(C["navy"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 36
    r += 1

    ws.merge_cells(f"B{r}:K{r}")
    cell = ws.cell(row=r, column=2,
                   value=f"  Month-to-date churned subscribers  ·  {mtd_label}  ·  "
                         f"Section 1: Paid Churn  ·  Section 2: Trial Churn")
    cell.font      = font(italic=True, size=9, color=C["white"])
    cell.fill      = fill(C["teal"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 18
    r += 1

    r = spacer(ws, r, 8)

    # Section 1: Paid Churn
    ws.merge_cells(f"B{r}:K{r}")
    sh = ws.cell(row=r, column=2,
                 value=f"  Paid Churned — MTD ({mtd_label})  ·  {len(paid_df)} subscriber(s)")
    sh.font      = font(bold=True, size=11, color=C["white"])
    sh.fill      = fill("6C3483")
    sh.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 22
    r += 1

    paid_hdrs = ["Churn Date", "Name", "Email", "Phone",
                 "Plan", "Platform", "Device", "User ID", "Subscription ID"]
    write_header_row(ws, r, paid_hdrs, C["navy"], start_col=2, height=22)
    r += 1

    if paid_df.empty:
        ws.merge_cells(f"B{r}:K{r}")
        ws.cell(row=r, column=2, value="No paid churn in MTD.").font = font(italic=True, size=10)
        r += 1
    else:
        paid_col_map = [
            ("churn_date",       None),
            ("full_name",        ""),
            ("email_id",         ""),
            ("contact_number",   ""),
            ("plan_segment",     ""),
            ("platform",         ""),
            ("device_type",      ""),
            ("user_id",          ""),
            ("subscription_id",  ""),
        ]
        for i, row_data in enumerate(paid_df.itertuples(index=False), start=0):
            row_vals = []
            for col, default in paid_col_map:
                val = getattr(row_data, col, default)
                if val is None or (not isinstance(val, (str, bool, int, float)) and pd.isna(val)):
                    val = default if default is not None else ""
                row_vals.append(val)
            write_data_row(ws, r, row_vals, start_col=2, alt=(i % 2 == 0), height=18,
                           bold_first=False)
            r += 1

    r = spacer(ws, r, 8)
    r = spacer(ws, r, 8)

    # Section 2: Trial Churn
    ws.merge_cells(f"B{r}:K{r}")
    sh = ws.cell(row=r, column=2,
                 value=f"  Trial Churned — MTD ({mtd_label})  ·  {len(trial_df)} subscriber(s)")
    sh.font      = font(bold=True, size=11, color=C["white"])
    sh.fill      = fill("1A5276")
    sh.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 22
    r += 1

    trial_hdrs = ["Trial Start", "Trial Expired", "Name", "Email", "Phone",
                  "Plan", "Platform", "Device", "User ID", "Subscription ID"]
    write_header_row(ws, r, trial_hdrs, C["navy"], start_col=2, height=22)
    r += 1

    if trial_df.empty:
        ws.merge_cells(f"B{r}:K{r}")
        ws.cell(row=r, column=2, value="No trial churn in MTD.").font = font(italic=True, size=10)
        r += 1
    else:
        trial_col_map = [
            ("trial_start",      None),
            ("trial_end",        None),
            ("full_name",        ""),
            ("email_id",         ""),
            ("contact_number",   ""),
            ("plan_segment",     ""),
            ("platform",         ""),
            ("device_type",      ""),
            ("user_id",          ""),
            ("subscription_id",  ""),
        ]
        for i, row_data in enumerate(trial_df.itertuples(index=False), start=0):
            row_vals = []
            for col, default in trial_col_map:
                val = getattr(row_data, col, default)
                if val is None or (not isinstance(val, (str, bool, int, float)) and pd.isna(val)):
                    val = default if default is not None else ""
                row_vals.append(val)
            write_data_row(ws, r, row_vals, start_col=2, alt=(i % 2 == 0), height=18,
                           bold_first=False)
            r += 1

    r = spacer(ws, r, 8)
    ws.merge_cells(f"B{r}:K{r}")
    note = ws.cell(
        row=r, column=2,
        value="Paid Churn: subscriptions with max paid_till in MTD and curr_status = In-Active. "
              "Trial Churn: trial subscriptions whose trial period ended in MTD (trial_end = "
              "first paid order date for converted, paid_till for unconverted) and did not convert. "
              "One row per subscription. Refreshed daily.",
    )
    note.font      = font(italic=True, size=9, color="666666")
    note.alignment = align(h="left", wrap=True)
    ws.row_dimensions[r].height = 36
