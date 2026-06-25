"""
data_loader.py
==============
Handles all data ingestion for the CEO Dashboard:
  1. load_from_excel()  — reads the mt_subs sheet from the source Excel file
  2. load_from_mysql()  — full reload from MySQL via the production query
  3. load_config()      — reads the ⚙️ Config sheet for alert thresholds
  4. write_raw_data()   — writes fresh MySQL data back to the source Excel file

Keeping data loading separate from KPI logic makes it easy to:
  - Switch data sources without changing KPI code
  - Test KPI logic against sample data without a DB connection
  - Add new data sources (e.g. S3, BigQuery) in one place
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import logging

log = logging.getLogger(__name__)

# ── Expected column names in the mt_subs sheet ────────────────────────────────
# Used for validation after loading to catch schema drift early.
REQUIRED_COLUMNS = [
    "order_identifier", "order_status", "user_id", "subscription_name",
    "duration", "subscription_amt", "discount", "total_amt",
    "trial_check", "subscribed_date", "paid_till", "curr_status", "subscription_id",
    "subscription_type", "device_type", "platform",
]

# ── Config sheet name ─────────────────────────────────────────────────────────
CONFIG_SHEET = "⚙️ Config"
DATA_SHEET   = "mt_subs"

# ── Default config values (used if Config sheet is missing a key) ─────────────
DEFAULT_CONFIG = {
    "churn_rate_warn"  : 0.05,   # 5%  — amber alert
    "churn_rate_bad"   : 0.08,   # 8%  — red alert
    "trial_conv_warn"  : 0.30,   # 30% — amber alert
    "trial_conv_bad"   : 0.20,   # 20% — red alert
    "free_conv_warn"   : 0.10,   # 10% — amber alert
}


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL SOURCE LOADER
# ══════════════════════════════════════════════════════════════════════════════

def load_from_excel(filepath: str) -> pd.DataFrame:
    """
    Load the mt_subs sheet from the source Excel file.

    Parameters
    ----------
    filepath : path to CEO_Subscription_Dashboard_ref.xlsx

    Returns
    -------
    pd.DataFrame with one row per order_identifier

    Raises
    ------
    FileNotFoundError  : if the file doesn't exist
    ValueError         : if required columns are missing (schema drift)
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Source file not found: {filepath}")

    log.info(f"Loading data from Excel: {filepath}")
    df = pd.read_excel(filepath, sheet_name=DATA_SHEET)

    # Strip whitespace from column names (common Excel issue)
    df.columns = df.columns.str.strip()

    # Rename legacy column name if present
    if "trial _check" in df.columns:
        df.rename(columns={"trial _check": "trial_check"}, inplace=True)
    if "plan_status" in df.columns and "curr_status" not in df.columns:
        df.rename(columns={"plan_status": "curr_status"}, inplace=True)

    # Validate required columns exist
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        log.warning(f"Missing columns (may be from older data version): {missing}")
        # Add missing columns with None so downstream code doesn't crash
        for col in missing:
            df[col] = None

    # Ensure date column is proper Python date
    df["subscribed_date"] = pd.to_datetime(df["subscribed_date"], errors="coerce").dt.date

    # Parse paid_till as Python date (may be NULL for some rows)
    if "paid_till" in df.columns:
        df["paid_till"] = pd.to_datetime(df["paid_till"], errors="coerce").dt.date

    # Ensure numeric columns
    for col in ["total_amt", "subscription_amt", "discount"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Derive mrr_amount if not present (SQL v2 adds this; older files may lack it)
    if "mrr_amount" not in df.columns:
        df["mrr_amount"] = df.apply(_derive_mrr, axis=1)

    # Derive plan_segment if not present
    if "plan_segment" not in df.columns:
        df["plan_segment"] = df.apply(_derive_plan_segment, axis=1)

    log.info(f"Loaded {len(df):,} rows, {df['user_id'].nunique():,} unique users")
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  MYSQL LOADER
# ══════════════════════════════════════════════════════════════════════════════

def load_from_mysql(host: str, port: int, user: str, password: str,
                    database: str, query: str) -> pd.DataFrame:
    """
    Execute the production SQL query against MySQL and return a DataFrame.

    Full reload on every call — no incremental logic needed at this scale.

    Parameters
    ----------
    host, port, user, password, database : MySQL connection params
    query : the SQL string to execute (from mileage_tracker_subscriptions.sql)

    Returns
    -------
    pd.DataFrame — same schema as load_from_excel()

    Requirements
    ------------
    pip install pymysql
    """
    try:
        import pymysql
    except ImportError:
        raise ImportError("pymysql not installed. Run: pip install pymysql")

    log.info(f"Connecting to MySQL: {host}:{port}/{database} as {user}")
    conn = pymysql.connect(
        host=host, port=port, user=user, password=password,
        database=database, charset="utf8mb4", connect_timeout=30,
    )
    try:
        log.info("Executing query ...")
        df = pd.read_sql(query, conn)
        log.info(f"Fetched {len(df):,} rows from MySQL")
    finally:
        conn.close()

    # Derive computed columns that the SQL v2 may not yet include
    # (safe to run even if columns already exist — will overwrite with same values)
    if "mrr_amount" not in df.columns:
        df["mrr_amount"] = df.apply(_derive_mrr, axis=1)
    if "plan_segment" not in df.columns:
        df["plan_segment"] = df.apply(_derive_plan_segment, axis=1)
    # Parse paid_till as Python date
    if "paid_till" in df.columns:
        df["paid_till"] = pd.to_datetime(df["paid_till"], errors="coerce").dt.date

    return df


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG LOADER
# ══════════════════════════════════════════════════════════════════════════════

def load_config(filepath: str) -> dict:
    """
    Read alert thresholds from the ⚙️ Config sheet in the dashboard workbook.

    Expected Config sheet layout (rows, columns A and B):
      Row 1 : header  (ignored)
      Row 2+: key | value

    Example:
      churn_rate_warn  | 0.05
      churn_rate_bad   | 0.08
      trial_conv_warn  | 0.30
      trial_conv_bad   | 0.20
      free_conv_warn   | 0.10

    Falls back to DEFAULT_CONFIG for any missing key.
    Returns an empty default config if the file or sheet doesn't exist.
    """
    config = DEFAULT_CONFIG.copy()

    if not os.path.exists(filepath):
        log.warning(f"Config file not found: {filepath}. Using defaults.")
        return config

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if CONFIG_SHEET not in wb.sheetnames:
            log.warning(f"'{CONFIG_SHEET}' sheet not found. Using default config.")
            return config

        ws = wb[CONFIG_SHEET]
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows and header rows
            if not row[0] or not row[1]:
                continue
            key = str(row[0]).strip()
            try:
                value = float(row[1])
                config[key] = value
                log.debug(f"Config: {key} = {value}")
            except (ValueError, TypeError):
                log.warning(f"Config row skipped — invalid value: {row}")

    except Exception as e:
        log.warning(f"Could not read config sheet: {e}. Using defaults.")

    return config


# ══════════════════════════════════════════════════════════════════════════════
#  WRITE RAW DATA BACK TO SOURCE EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def write_raw_data(df: pd.DataFrame, filepath: str):
    """
    Write the freshly loaded DataFrame back to the mt_subs sheet
    in the source Excel file.

    Called after load_from_mysql() to persist the data locally,
    so the dashboard can be rebuilt without a DB connection if needed.

    Preserves all other sheets (Config, Documentation, etc.) in the workbook.

    Parameters
    ----------
    df       : DataFrame to write
    filepath : path to the Excel workbook
    """
    log.info(f"Writing {len(df):,} rows to {filepath} [{DATA_SHEET}]")

    # Load existing workbook if it exists, else create new
    if os.path.exists(filepath):
        try:
            wb = openpyxl.load_workbook(filepath)
            if DATA_SHEET in wb.sheetnames:
                del wb[DATA_SHEET]
        except Exception as e:
            log.warning(f"Could not open existing ref file ({e}); creating fresh workbook.")
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
    else:
        wb = openpyxl.Workbook()
        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet(DATA_SHEET)

    # Write header row with styling
    header_fill = PatternFill("solid", fgColor="1B2A4A")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center")

    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        # Set column width based on header length
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max(14, len(col_name) + 2)

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    # Write data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            # Convert NaT/NaN to None for clean Excel output
            if pd.isna(val) if not isinstance(val, (str, bool)) else False:
                val = None
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Apply auto-filter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}1"

    wb.save(filepath)
    log.info(f"Source data written to {filepath}")


# ══════════════════════════════════════════════════════════════════════════════
#  DERIVATION HELPERS (used when SQL columns are absent)
# ══════════════════════════════════════════════════════════════════════════════

def _derive_mrr(row) -> float:
    """
    Derive mrr_amount from total_amt + duration.
    Yearly plans: total_amt / 12
    Monthly plans: total_amt as-is
    Free/Trial: 0
    """
    amt = float(row.get("total_amt", 0) or 0)
    if amt <= 0:
        return 0.0
    duration = str(row.get("duration", "")).lower()
    if "year" in duration:
        return round(amt / 12, 2)
    return round(amt, 2)


def _derive_plan_segment(row) -> str:
    """
    Derive plan_segment from subscription_name + duration.
    Free → Free Monthly (Free has no yearly option)
    Others → "{Name} {Duration}"
    """
    name = str(row.get("subscription_name", "")).strip()
    dur  = str(row.get("duration", "Monthly")).strip()
    if name == "Free":
        return "Free Monthly"
    return f"{name} {dur}"


# ══════════════════════════════════════════════════════════════════════════════
#  SUBSCRIPTION STATE SNAPSHOTS
# ══════════════════════════════════════════════════════════════════════════════

SNAPSHOT_SHEET = "Monthly Active Snapshots"
SNAPSHOT_COLS  = ["snapshot_date", "subscription_id", "user_id", "plan_segment", "paid_till"]


def read_active_snapshot(filepath: str) -> pd.DataFrame:
    """Read the Monthly Active Snapshots sheet from an existing dashboard file.
    Returns empty DataFrame with correct columns if file/sheet doesn't exist."""
    empty = pd.DataFrame(columns=SNAPSHOT_COLS)
    if not os.path.exists(filepath):
        return empty
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if SNAPSHOT_SHEET not in wb.sheetnames:
            wb.close()
            return empty
        ws = wb[SNAPSHOT_SHEET]
        rows = list(ws.values)
        wb.close()
        if len(rows) <= 1:
            return empty
        headers = [str(h) for h in rows[0]]
        data = rows[1:]
        df = pd.DataFrame(data, columns=headers)
        df["paid_till"] = pd.to_datetime(df["paid_till"], errors="coerce").dt.date
        return df
    except Exception as e:
        log.warning(f"Could not read snapshot sheet: {e}")
        return empty


def update_active_snapshot(df: pd.DataFrame, snapshot_date) -> pd.DataFrame:
    """Regenerate the full Monthly Active Snapshot from scratch on every run.

    Generates one row per active subscription for every month-start from
    Jan 1 of the current year through snapshot_date (inclusive).

    Runs fresh each time — no idempotency check. This guarantees the snapshot
    sheet is always consistent with compute_active_by_plan and
    compute_active_subscribers, which also recompute from scratch on every run.

    snapshot_date: a date object (typically MTD_START = first of current month).

    Active logic (mirrors _active_count_on() in kpi_engine.py):
      subscribed_date <= cutoff  (no future orders inflate historical counts)
      Case A: end_date IS NULL or paid_till > end_date → active if paid_till > cutoff
      Case B: end_date IS NOT NULL and paid_till <= end_date → active if end_date > cutoff
    """
    from datetime import date, timedelta

    has_paid_till = "paid_till" in df.columns and df["paid_till"].notna().any()
    if not has_paid_till:
        log.warning("paid_till not available — skipping snapshot generation")
        return pd.DataFrame(columns=SNAPSHOT_COLS)

    year_start   = date(snapshot_date.year, 1, 1)
    # Build month-starts: first of each month from Jan 1 through snapshot_date
    month_dates  = []
    d = year_start
    while d <= snapshot_date:
        month_dates.append(d)
        # advance to first of next month
        if d.month == 12:
            d = date(d.year + 1, 1, 1)
        else:
            d = date(d.year, d.month + 1, 1)

    records = []
    for snap_d in month_dates:
        cutoff = snap_d - timedelta(days=1)

        df_eligible = df[df["subscribed_date"] <= cutoff].copy()
        if df_eligible.empty:
            continue

        agg_dict = dict(
            user_id      = ("user_id",      "first"),
            plan_segment = ("plan_segment", "last"),
            paid_till    = ("paid_till",    "max"),
        )
        if "end_date" in df_eligible.columns:
            agg_dict["end_date"] = ("end_date", "max")

        sub_snap = df_eligible.groupby("subscription_id", as_index=False).agg(**agg_dict)

        if "end_date" in sub_snap.columns:
            case_b = sub_snap["end_date"].notna() & (sub_snap["paid_till"] <= sub_snap["end_date"])
            active_anchor = sub_snap["paid_till"].copy()
            active_anchor[case_b] = sub_snap.loc[case_b, "end_date"]
            sub_snap = sub_snap[active_anchor > cutoff].copy()
        else:
            sub_snap = sub_snap[sub_snap["paid_till"] > cutoff].copy()

        sub_snap["snapshot_date"] = str(snap_d)
        records.append(sub_snap)
        log.info(f"  Snapshot {snap_d}: {len(sub_snap):,} active subscriptions")

    if not records:
        log.warning("No snapshot records generated")
        return pd.DataFrame(columns=SNAPSHOT_COLS)

    result = pd.concat(records, ignore_index=True)
    cols   = [c for c in SNAPSHOT_COLS if c in result.columns]
    log.info(f"  Total snapshot rows: {len(result):,} across {len(month_dates)} month(s)")
    return result[cols]


def write_snapshot_sheet(wb, snapshot_df: pd.DataFrame):
    """Write the Monthly Active Snapshots sheet into an openpyxl Workbook.
    Called after all other sheets are built so it appears at the end."""
    if SNAPSHOT_SHEET in wb.sheetnames:
        del wb[SNAPSHOT_SHEET]
    ws = wb.create_sheet(SNAPSHOT_SHEET)
    ws.sheet_view.showGridLines = True
    ws.sheet_state = "visible"

    header_fill_color = "1B2A4A"
    if snapshot_df.empty:
        for ci, col in enumerate(SNAPSHOT_COLS, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor=header_fill_color)
        return

    cols = [c for c in SNAPSHOT_COLS if c in snapshot_df.columns]
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor=header_fill_color)

    for ri, row_data in enumerate(snapshot_df[cols].itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, start=1):
            if not isinstance(val, (str, bool)) and pd.isna(val):
                val = None
            ws.cell(row=ri, column=ci, value=val)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols))}1"
