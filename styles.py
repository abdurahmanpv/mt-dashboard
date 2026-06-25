"""
styles.py
=========
Shared style constants and helper functions for building the Excel dashboard.

All colours, fonts, borders and cell-writing utilities live here so that:
  - Every sheet uses a consistent visual language
  - You can update the entire dashboard's look by editing this one file
  - Sheet builder files stay focused on layout, not style boilerplate
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── COLOUR PALETTE ────────────────────────────────────────────────────────────
# All hex colours without the leading #
C = {
    # Primary brand colours
    "navy"      : "1B2A4A",
    "teal"      : "0E7C7B",
    "gold"      : "F5A623",

    # Status colours
    "green"     : "1E8449",
    "red"       : "C0392B",
    "amber"     : "D68910",

    # Background shades
    "white"     : "FFFFFF",
    "light_bg"  : "F0F4F8",   # alternating row light shade
    "card_bg"   : "EBF5FB",   # KPI card background
    "note_bg"   : "FDFEFE",   # documentation note background

    # Alert backgrounds / foregrounds
    "good_bg"   : "D5F5E3",
    "good_fg"   : "145A32",
    "warn_bg"   : "FEF9E7",
    "warn_fg"   : "784212",
    "bad_bg"    : "FADBD8",
    "bad_fg"    : "922B21",
    "info_bg"   : "EBF5FB",
    "info_fg"   : "1B4F72",

    # Text
    "text_dk"   : "1A1A2E",
    "text_md"   : "34495E",
    "text_lt"   : "7F8C8D",

    # Section headers
    "sub_hdr"   : "2C3E50",
    "border"    : "BDC3C7",
}

# ── SIGNAL LEVEL → COLOURS ───────────────────────────────────────────────────
SIGNAL_COLORS = {
    "GOOD" : (C["good_bg"], C["good_fg"]),
    "WARN" : (C["warn_bg"], C["warn_fg"]),
    "BAD"  : (C["bad_bg"],  C["bad_fg"]),
    "INFO" : (C["info_bg"], C["info_fg"]),
}


# ══════════════════════════════════════════════════════════════════════════════
#  STYLE FACTORIES
#  Functions that return openpyxl style objects.
#  Call these inline rather than storing style objects in variables
#  (openpyxl styles are mutable and can bleed between cells if reused).
# ══════════════════════════════════════════════════════════════════════════════

def fill(hex_color: str) -> PatternFill:
    """Solid fill with the given hex colour."""
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, size=11, color=None, italic=False, name="Calibri") -> Font:
    """Calibri font with optional bold/italic/size/colour."""
    return Font(
        name=name, bold=bold, size=size,
        color=color or C["text_dk"],
        italic=italic,
    )


def border(style="thin", color=None) -> Border:
    """Uniform border on all 4 sides."""
    s = Side(border_style=style, color=color or C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def align(h="center", v="center", wrap=False) -> Alignment:
    """Alignment helper."""
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ══════════════════════════════════════════════════════════════════════════════
#  CELL WRITER HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def write_cell(ws, row: int, col: int, value,
               bold=False, size=11, color=None, italic=False,
               bg=None, border_style=None,
               h_align="center", v_align="center", wrap=False,
               number_format=None) -> object:
    """
    Write a value to a cell and apply all styles in one call.

    Returns the cell object so you can do further customisation if needed.
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = font(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = align(h=h_align, v=v_align, wrap=wrap)
    if bg:           cell.fill   = fill(bg)
    if border_style: cell.border = border(border_style)
    if number_format: cell.number_format = number_format
    return cell


def write_header_row(ws, row: int, labels: list, bg: str,
                     font_color: str = "FFFFFF",
                     start_col: int = 1,
                     size: int = 10,
                     height: int = 18):
    """Write a styled header row across multiple columns."""
    ws.row_dimensions[row].height = height
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=label)
        cell.font      = font(bold=True, size=size, color=font_color)
        cell.fill      = fill(bg)
        cell.alignment = align()
        cell.border    = border()


def write_data_row(ws, row: int, values: list,
                   start_col: int = 1,
                   alt: bool = False,
                   height: int = 17,
                   bold_first: bool = True,
                   first_align: str = "left"):
    """
    Write a data row with alternating background and optional bold first cell.

    alt=True  → light_bg background
    alt=False → white background
    """
    bg = C["light_bg"] if alt else C["white"]
    ws.row_dimensions[row].height = height
    for i, val in enumerate(values):
        h = first_align if i == 0 else "center"
        bold = bold_first and i == 0
        if isinstance(val, tuple) and len(val) == 2:
            cell = ws.cell(row=row, column=start_col + i, value=val[0])
            if val[1]:
                cell.number_format = val[1]
        else:
            cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.font      = font(bold=bold, size=10)
        cell.fill      = fill(bg)
        cell.border    = border("thin")
        cell.alignment = align(h=h, v="center", wrap=True)


def write_section_header(ws, row: int, col_start: int, col_end: int,
                         title: str, bg: str = None,
                         height: int = 22) -> int:
    """
    Write a full-width section header banner and return the next row number.
    Merges cells from col_start to col_end.
    """
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=f"  {title}")
    cell.font      = font(bold=True, size=12, color=C["white"])
    cell.fill      = fill(bg or C["sub_hdr"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[row].height = height
    return row + 1


def spacer(ws, row: int, height: int = 8) -> int:
    """Add a blank spacer row and return the next row number."""
    ws.row_dimensions[row].height = height
    return row + 1


def set_col_widths(ws, widths: list, start_col: int = 1):
    """Set column widths from a list, starting at start_col."""
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
#  FORMATTING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def fmt_num(value, prefix="", suffix="", decimals=0) -> str:
    """Format a number with optional prefix/suffix. Returns '–' for None."""
    if value is None:
        return "–"
    try:
        if decimals == 0:
            return f"{prefix}{int(value):,}{suffix}"
        return f"{prefix}{value:,.{decimals}f}{suffix}"
    except (TypeError, ValueError):
        return str(value)


def fmt_pct(value, decimals=1) -> str:
    """Format a float as percentage. Returns '–' for None."""
    if value is None:
        return "–"
    return f"{value * 100:.{decimals}f}%"


def fmt_currency(value, decimals=2) -> str:
    """Format a float as USD currency. Returns '–' for None."""
    if value is None:
        return "–"
    return f"${value:,.{decimals}f}"


def nv(x, decimals=0):
    """Return (raw_value, Excel_number_format) for numeric counts, or '–' if None."""
    if x is None:
        return "–"
    fmt = "#,##0" if decimals == 0 else "#,##0." + "0" * decimals
    return (int(x) if decimals == 0 else round(float(x), decimals), fmt)


def pv(x, decimals=1):
    """Return (raw_value, Excel_number_format) for percentages, or '–' if None."""
    if x is None:
        return "–"
    return (round(float(x), decimals + 2), "0." + "0" * decimals + "%")


def cv(x, decimals=2):
    """Return (raw_value, Excel_number_format) for currency, or '–' if None."""
    if x is None:
        return "–"
    return (round(float(x), decimals), '"$"#,##0.' + "0" * decimals)


def variance_str(var_dict: dict, is_pct: bool = False) -> str:
    """
    Format a variance dict {abs, pct, direction} as a display string.
    e.g. "▲ 12 (+8.3%)" or "▼ 3 (-2.1%)"
    """
    if not var_dict:
        return "–"
    arrow = "▲" if var_dict["direction"] == "up" else ("▼" if var_dict["direction"] == "down" else "→")
    abs_val = var_dict["abs"]
    pct_val = var_dict.get("pct")

    if is_pct:
        abs_str = fmt_pct(abs(abs_val) / 100)
    else:
        abs_str = fmt_num(abs(abs_val))

    if pct_val is not None:
        return f"{arrow} {abs_str} ({'+' if abs_val >= 0 else ''}{pct_val:.1f}%)"
    return f"{arrow} {abs_str}"
