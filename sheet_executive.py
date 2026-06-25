"""
sheet_executive.py
==================
Builds the Executive Dashboard sheet.

Layout:
  1. Title banner
  2. KPI cards (3 rows × 4 cards):
       Row 1: Unique Users | Unique Subscriptions | Total Transactions | Active Subs
       Row 2: New Subs     | Trials               | Renewals       | Trial Pipeline
       Row 3: Cash Revenue | MRR                  | Paid Churn Rate | Trial→Paid Conv
  3. Full metrics table (all KPIs × all time windows, with cell comments on health metrics)
  4. MoM variance summary
  5. Decision signals
  6. 12-month trend summary table
"""

import openpyxl
from openpyxl.comments import Comment
from styles import (C, fill, font, border, align,
                    write_header_row, write_data_row, write_section_header,
                    spacer, set_col_widths, fmt_num, fmt_pct, fmt_currency,
                    variance_str, SIGNAL_COLORS,
                    nv, pv, cv)


def _pct_chg(current, previous):
    """Return a formatted % change string with arrow, e.g. '▲ +12.3% vs prev MTD'"""
    try:
        if previous is None or previous == 0:
            return "vs prev MTD: –"
        delta = (current - previous) / abs(previous) * 100
        arrow = "▲" if delta > 0 else ("▼" if delta < 0 else "→")
        sign  = "+" if delta > 0 else ""
        return f"{arrow} {sign}{delta:.1f}%  vs prev MTD"
    except Exception:
        return "vs prev MTD: –"


def build_executive_dashboard(wb: openpyxl.Workbook, kpi: dict):
    """
    Build the Executive Dashboard sheet.

    Parameters
    ----------
    wb  : openpyxl Workbook
    kpi : full KPI dict from kpi_engine.compute_all_kpis()
    """
    # Remove existing sheet if rebuilding
    if "Executive Dashboard" in wb.sheetnames:
        del wb["Executive Dashboard"]

    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_view.showGridLines = False

    # Column layout: [margin, label, yesterday, mtd, prev_full, prev_same, ytd, lifetime, margin]
    # Col indices:       1      2       3         4     5          6          7    8          9
    set_col_widths(ws, [0.8, 32, 16, 16, 18, 20, 16, 16, 0.8])

    r = 1

    # ── TITLE BANNER ──────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 6
    r += 1

    ws.merge_cells(f"B{r}:I{r}")
    cell = ws.cell(row=r, column=2,
                   value="  CEO SUBSCRIPTION & REVENUE DASHBOARD")
    cell.font      = font(bold=True, size=20, color=C["white"])
    cell.fill      = fill(C["navy"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 44
    r += 1

    ws.merge_cells(f"B{r}:I{r}")
    days = kpi["days_elapsed_mtd"]
    tz   = kpi.get("tz_label", "Pacific Time (PT)")
    stale_tag = (f"  ⚠ DATA STALE — source is {kpi['data_lag_days']}d behind "
                 f"(run daily_refresh.py to fix)  ·  "
                 if kpi.get("data_stale") else "  ")
    cell = ws.cell(row=r, column=2,
                   value=f"{stale_tag}Data through {kpi['as_of']} ({tz})  ·  "
                         f"MTD = {kpi['mtd_start']} → {kpi['as_of']} ({days} days)  ·  "
                         f"Prev Month = {kpi['prev_month_start']} → {kpi['prev_month_end']}  ·  "
                         f"Same-period = {kpi['prev_month_start']} → {kpi['prev_same_end']}  ·  "
                         f"Generated: {kpi.get('generated_on', kpi['as_of'])}")
    cell.font      = font(italic=True, size=10, color=C["white"])
    cell.fill      = fill(C["amber"] if kpi.get("data_stale") else C["teal"])
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 20
    r += 1

    # ── MTD period header ────────────────────────────────────────────────────
    ws.merge_cells(f"B{r}:I{r}")
    cell = ws.cell(row=r, column=2,
                   value=f"  MTD PERIOD:  {kpi['lbl_mtd']}  ·  "
                         f"Previous MTD for comparison:  {kpi['lbl_prev_same']}  "
                         f"  (same-period last month)")
    cell.font      = font(bold=True, size=11, color=C["navy"])
    cell.fill      = fill("EBF5FB")
    cell.alignment = align(h="left", v="center")
    ws.row_dimensions[r].height = 24
    r += 1

    r = spacer(ws, r, 10)

    # ── KPI CARDS ─────────────────────────────────────────────────────────────
    # 3 rows of 4 cards each — each card spans 2 columns (B-C, D-E, F-G, H-I)
    # Card structure: title (row 1), big number (row 2), sub-label (row 3),
    #                 accent bar (row 4)

    sc = kpi["subscription_counts"]
    st = kpi["subscription_types"]
    rev = kpi["revenue"]
    canc = kpi.get("cancellations", {w: {"count": 0, "refund_amount": 0.0} for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]})
    act = kpi["active"]
    pchurn = kpi["paid_churn"]
    t_conv = kpi["trial_conversion"]
    t_pipe = kpi["trial_pipeline"]
    f_conv = kpi["free_conversion"]

    as_of    = kpi["as_of"]          # e.g. "2026-06-07"

    # Row 1 cards: volume overview
    card_rows = [
        [
            ("UNIQUE USERS",         nv(sc["unique_users"]["mtd"]),
             _pct_chg(sc["unique_users"]["mtd"], sc["unique_users"]["prev_same"]),
             f"Yesterday: {fmt_num(sc['unique_users']['yesterday'])}   ·   "
             f"Lifetime: {fmt_num(sc['unique_users']['lifetime'])}",   C["navy"]),
            ("UNIQUE SUBSCRIPTIONS", nv(sc["unique_subs"]["mtd"]),
             _pct_chg(sc["unique_subs"]["mtd"], sc["unique_subs"]["prev_same"]),
             f"Yesterday: {fmt_num(sc['unique_subs']['yesterday'])}   ·   "
             f"Lifetime: {fmt_num(sc['unique_subs']['lifetime'])}",    C["teal"],
             "Number of distinct subscription plans in this period. "
             "One customer can hold more than one subscription at a time."),
            ("TOTAL TRANSACTIONS",    nv(sc["total_transactions"]["mtd"]),
             _pct_chg(sc["total_transactions"]["mtd"], sc["total_transactions"]["prev_same"]),
             f"Yesterday: {fmt_num(sc['total_transactions']['yesterday'])}   ·   "
             f"Lifetime: {fmt_num(sc['total_transactions']['lifetime'])}",  "16537A",
             "Total transactions in this period — trial starts, new subscriptions, and renewals. "
             "A customer who renews each month creates one transaction per month."),
            ("ACTIVE SUBSCRIBERS",   nv(act["active_now"]),
             "Point-in-time  ·  as of end of yesterday",
             f"At month start: {fmt_num(act['active_at_month_start'])}   ·   "
             f"Paid: {fmt_num(pchurn['mtd']['denominator'])}  Free: {fmt_num(act['active_at_month_start'] - pchurn['mtd']['denominator'])}",
             "6C3483",
             "Active subscriptions as of end of yesterday. "
             "Includes both paid and free plan subscribers. "
             "Note: This is a point-in-time snapshot, NOT an MTD flow metric. % change is not shown."),
        ],
        [
            ("NEW SUBSCRIPTIONS",    nv(st["new_subs"]["mtd"]),
             _pct_chg(st["new_subs"]["mtd"], st["new_subs"]["prev_same"]),
             f"Yesterday: {fmt_num(st['new_subs']['yesterday'])}   ·   "
             f"Lifetime: {fmt_num(st['new_subs']['lifetime'])}",        C["navy"]),
            ("TRIALS STARTED",       nv(st["trials"]["mtd"]),
             _pct_chg(st["trials"]["mtd"], st["trials"]["prev_same"]),
             f"Yesterday: {fmt_num(st['trials']['yesterday'])}   ·   "
             f"Lifetime: {fmt_num(st['trials']['lifetime'])}",          C["teal"]),
            ("RENEWALS",             nv(st["renewals"]["mtd"]),
             _pct_chg(st["renewals"]["mtd"], st["renewals"]["prev_same"]),
             f"Yesterday: {fmt_num(st['renewals']['yesterday'])}   ·   "
             f"Prev Month: {fmt_num(st['renewals']['prev_full'])}",     "16537A"),
            ("TRIAL PIPELINE",       nv(t_pipe["count"]),
             "Open trials",
             "Unconverted trials still within their trial period",
             "6C3483",
             "Customers who started a free trial and have not yet converted to a paid subscription. "
             "A trial is counted here if paid_till >= today (access has not yet expired). "
             "Note: This is a point-in-time snapshot, NOT an MTD flow metric. % change is not shown."),
        ],
        [
            ("CASH REVENUE",         cv(rev["mtd"]["net_cash"]),
             _pct_chg(rev["mtd"]["net_cash"], rev["prev_same"]["net_cash"]),
             f"Gross: {fmt_currency(rev['mtd']['cash'])}   ·   "
             f"Refunds: {fmt_currency(rev['mtd']['refunds'])}",         C["navy"]),
            ("MRR",                  cv(rev["mtd"]["mrr"]),
             _pct_chg(rev["mtd"]["mrr"], rev["prev_same"]["mrr"]),
             f"Yesterday: {fmt_currency(rev['yesterday']['mrr'])}   ·   "
             f"YTD MRR: {fmt_currency(rev['ytd']['mrr'])}",             C["teal"]),
            ("PAID CHURN RATE",      pv(pchurn["mtd"]["rate"]),
             _pct_chg(pchurn["mtd"]["rate"] * 100,
                      pchurn["prev_full"]["rate"] * 100),
             f"Churn events: {fmt_num(pchurn['mtd']['events'])}   ·   "
             f"Paid active at start: {fmt_num(pchurn['mtd']['denominator'])}",
             "C0392B",
             "Share of paid subscribers who did not renew during this period. "
             "Formula: Subscribers who churned ÷ Active paid subscribers at period start. "
             "Excludes free and trial plans. Lower is better. "
             "% change compares MTD vs full previous month (prev_same not available for churn)."),
            ("TRIAL→PAID CONV",
             pv(t_conv["mtd"]["rate"]) if t_conv["mtd"]["rate"] is not None else "–",
             _pct_chg((t_conv["mtd"]["rate"] or 0) * 100,
                      (t_conv["prev_full"]["rate"] or 0) * 100),
             f"Trials ended: {fmt_num(t_conv['mtd']['total'])}   ·   "
             f"Converted: {fmt_num(t_conv['mtd']['converted'])}",       "1E8449",
             "Of all 7-day trials that ended this month, the share that converted to a paid subscription. "
             "Formula: Paid conversions ÷ Trials whose 7-day window ended this month. "
             "% change compares MTD vs full previous month (prev_same not available for trial conv)."),
        ],
    ]

    card_col_starts = [2, 4, 6, 8]

    for row_group in card_rows:
        # 5 rows per card group: spacer + title + value + period + sub + bar
        ws.row_dimensions[r].height = 6;   r_title = r + 1
        ws.row_dimensions[r+1].height = 14
        ws.row_dimensions[r+2].height = 34
        ws.row_dimensions[r+3].height = 14
        ws.row_dimensions[r+4].height = 18
        ws.row_dimensions[r+5].height = 10

        for col_start, card_data in zip(card_col_starts, row_group):
            title, value, period, sub, color = card_data[:5]
            card_comment = card_data[5] if len(card_data) > 5 else None
            for row_offset, fill_r in enumerate(range(r+1, r+6)):
                for c in range(col_start, col_start + 2):
                    ws.cell(row=fill_r, column=c).fill = fill(C["card_bg"])

            # Title
            ws.merge_cells(start_row=r+1, start_column=col_start,
                           end_row=r+1, end_column=col_start+1)
            cell = ws.cell(row=r+1, column=col_start, value=title)
            cell.font      = font(bold=True, size=9, color=color)
            cell.fill      = fill(C["card_bg"])
            cell.alignment = align(h="center", v="center")

            # Value (big number — represents MTD)
            ws.merge_cells(start_row=r+2, start_column=col_start,
                           end_row=r+2, end_column=col_start+1)
            if isinstance(value, tuple) and len(value) == 2:
                cell = ws.cell(row=r+2, column=col_start, value=value[0])
                cell.number_format = value[1]
            else:
                cell = ws.cell(row=r+2, column=col_start, value=value)
            cell.font      = font(bold=True, size=20, color=color)
            cell.fill      = fill(C["card_bg"])
            cell.alignment = align(h="center", v="center")

            # Period label (e.g. "MTD · Jun 1–7")
            ws.merge_cells(start_row=r+3, start_column=col_start,
                           end_row=r+3, end_column=col_start+1)
            cell = ws.cell(row=r+3, column=col_start, value=period)
            cell.font      = font(size=8, color=color, italic=True)
            cell.fill      = fill(C["card_bg"])
            cell.alignment = align(h="center", v="center")

            # Sub-label (Yesterday comparison)
            ws.merge_cells(start_row=r+4, start_column=col_start,
                           end_row=r+4, end_column=col_start+1)
            cell = ws.cell(row=r+4, column=col_start, value=sub)
            cell.font      = font(size=8, color=C["text_md"], italic=True)
            cell.fill      = fill(C["card_bg"])
            cell.alignment = align(h="center", v="center", wrap=True)

            # Accent bar
            ws.merge_cells(start_row=r+5, start_column=col_start,
                           end_row=r+5, end_column=col_start+1)
            cell = ws.cell(row=r+5, column=col_start, value="")
            cell.fill = fill(color)

            # Optional cell comment on the title cell
            if card_comment:
                try:
                    ws.cell(row=r+1, column=col_start).comment = Comment(card_comment, "Dashboard")
                except Exception:
                    pass

        r += 6

    r = spacer(ws, r, 10)

    # ── FULL METRICS TABLE ────────────────────────────────────────────────────
    r = write_section_header(ws, r, 2, 8, "PERFORMANCE METRICS — ALL TIME WINDOWS")

    # Column headers with explicit date ranges (single-line — no embedded newlines;
    # literal \n in cell values corrupts sheet XML when Excel opens the file)
    hdrs = [
        "Metric",
        f"Yesterday ({kpi['lbl_yesterday']})",
        f"MTD ({kpi['lbl_mtd']})",
        f"Prev Month Full ({kpi['lbl_prev_full']})",
        f"Same Period Last Mo. ({kpi['lbl_prev_same']})",
        f"YTD ({kpi['lbl_ytd']})",
        f"Lifetime (All → {kpi['lbl_yesterday']})",
    ]
    write_header_row(ws, r, hdrs, C["navy"], start_col=2, height=36)
    r += 1

    def _row(label, vals, alt=False, bold_first=True):
        write_data_row(ws, r, [label] + list(vals),
                       start_col=2, alt=alt, height=18,
                       bold_first=bold_first, first_align="left")

    def _note(row_num, col_num, text):
        """Attach a cell comment/note to the given cell."""
        try:
            ws.cell(row=row_num, column=col_num).comment = Comment(text, "Dashboard")
        except Exception:
            pass  # never let comment failures break the build

    # Build metrics rows — (label, [yesterday, mtd, prev_full, prev_same, ytd, lifetime], optional note)
    metrics = [
        # (label, vals, note_text_or_None)
        ("Unique Users",
         [nv(sc["unique_users"][w]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         "How many individual customers placed an order in this period, "
         "regardless of how many plans they hold."),
        ("Unique Subscriptions",
         [nv(sc["unique_subs"][w]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         "How many distinct subscription plans were created in this period. "
         "One customer can hold more than one subscription at a time."),
        ("Total Transactions",
         [nv(sc["total_transactions"][w]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         "Total transactions in this period — trial starts, new subscriptions, and renewals. "
         "A customer renewing monthly creates one transaction per month."),
        ("── By Type ──────", ["","","","","",""], None),
        ("  Trials",
         [nv(st["trials"][w]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         None),
        ("  New Subscriptions",
         [nv(st["new_subs"][w]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         None),
        ("  Renewals",
         [nv(st["renewals"][w]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         None),
        ("── By Plan Type ──", ["","","","","",""], None),
        ("  Paid",
         [nv(st["paid"][w]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         None),
        ("  Free",
         [nv(st["free"][w]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         None),
        ("── Revenue ───────", ["","","","","",""], None),
        ("  Gross Cash ($)",
         [cv(rev[w]["cash"]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         "Actual cash collected from Confirmed and Completed paid orders only."),
        ("  Refunds ($)",
         [cv(rev[w]["refunds"]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         "Total refund_amount on Cancelled/Refunded orders in this window."),
        ("  Net Cash ($)",
         [cv(rev[w]["net_cash"]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         "Gross Cash minus Refunds."),
        ("  MRR ($)",
         [cv(rev[w]["mrr"]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         "Accrual-based MRR: monthly plan orders in window + 1/12 of yearly orders in 12-month recognition window."),
        ("── Cancellations ─", ["","","","","",""], None),
        ("  Cancelled Orders",
         [nv(canc[w]["count"]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         "Count of orders with order_status = Cancelled or Refund."),
        ("  Total Refunded ($)",
         [cv(canc[w]["refund_amount"]) for w in ["yesterday","mtd","prev_full","prev_same","ytd","lifetime"]],
         "Sum of SBO_RefundAmount on cancelled/refunded orders."),
    ]

    for i, row_data in enumerate(metrics):
        label, vals, note_text = row_data
        # Style separator rows differently
        is_separator = label.startswith("──")
        if is_separator:
            ws.row_dimensions[r].height = 14
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            cell = ws.cell(row=r, column=2, value=label)
            cell.font = font(bold=True, size=9, color=C["text_lt"], italic=True)
            cell.fill = fill(C["mid_bg"] if "mid_bg" in C else "D6E4F0")
            cell.alignment = align(h="left", v="center")
        else:
            write_data_row(ws, r, [label] + vals,
                           start_col=2, alt=(i % 2 == 0), height=18,
                           bold_first=not label.startswith(" "),
                           first_align="left")
            if note_text:
                _note(r, 2, note_text)
        r += 1

    r = spacer(ws, r, 10)

    # ── MOM VARIANCE SUMMARY ──────────────────────────────────────────────────
    r = write_section_header(ws, r, 2, 8, "MoM VARIANCE SUMMARY")

    write_header_row(ws, r,
                     ["Metric", "MTD", "vs Full Prev Month", "vs Same Days Last Month"],
                     C["navy"], start_col=2, height=18)
    r += 1

    variance_rows = [
        ("New Subscriptions",
         st["new_subs"]["mtd"], st["new_subs"]["prev_full"], st["new_subs"]["prev_same"], False),
        ("Trials",
         st["trials"]["mtd"], st["trials"]["prev_full"], st["trials"]["prev_same"], False),
        ("Renewals",
         st["renewals"]["mtd"], st["renewals"]["prev_full"], st["renewals"]["prev_same"], False),
        ("Paid Subscribers",
         st["paid"]["mtd"], st["paid"]["prev_full"], st["paid"]["prev_same"], False),
        ("Gross Cash ($)",
         rev["mtd"]["cash"], rev["prev_full"]["cash"], rev["prev_same"]["cash"], False),
        ("Net Cash ($)",
         rev["mtd"]["net_cash"], rev["prev_full"]["net_cash"], rev["prev_same"]["net_cash"], False),
        ("MRR ($)",
         rev["mtd"]["mrr"], rev["prev_full"]["mrr"], rev["prev_same"]["mrr"], False),
    ]

    for i, (label, mtd_v, pf_v, ps_v, is_pct) in enumerate(variance_rows):
        ws.row_dimensions[r].height = 18
        v_full = _compute_var(mtd_v, pf_v)
        v_same = _compute_var(mtd_v, ps_v) if ps_v is not None else None
        fmt_val = pv(mtd_v) if is_pct else (cv(mtd_v) if isinstance(mtd_v, float) else nv(mtd_v))
        write_data_row(ws, r, [
            label,
            fmt_val,
            variance_str(v_full, is_pct) if v_full else "–",
            variance_str(v_same, is_pct) if v_same else "–",
        ], start_col=2, alt=(i % 2 == 0), height=18)
        r += 1

    r = spacer(ws, r, 10)

    # ── 12-MONTH TREND TABLE ──────────────────────────────────────────────────
    r = write_section_header(ws, r, 2, 8, "12-MONTH TREND (Last 12 Full Months + Current Month MTD)")

    trend_hdrs = ["Month", "New Subs", "Trials", "Renewals",
                  "Paid", "Cash Rev ($)", "MRR ($)", "Trial Conv%",
                  "Total (New+Renewal)", "Paid New Subs"]
    write_header_row(ws, r, trend_hdrs, C["teal"], start_col=2, height=18)
    trend_hdr_row  = r
    r += 1
    trend_data_start = r

    for i, month in enumerate(kpi["monthly_trend"]):
        bg = C["light_bg"] if i % 2 == 0 else C["white"]
        ws.row_dimensions[r].height = 17
        # Write raw numbers so the chart can reference them;
        # apply Excel number formats for display.
        values = [
            month["month_label"],   # col B (text)
            month["new_subs"],      # col C
            month["trials"],        # col D
            month["renewals"],      # col E
            month["paid"],          # col F
            month["cash_revenue"],  # col G
            month["mrr_revenue"],   # col H
            month["trial_conv_rate"] if month["trial_conv_rate"] is not None else 0,  # col I
            month["total_subs"],    # col J: New+Renewals (excludes trials)
            month["paid_new_subs"], # col K: paid new subscriptions only
        ]
        number_fmts = [None, "#,##0", "#,##0", "#,##0", "#,##0",
                       '"$"#,##0.00', '"$"#,##0.00', "0.0%", "#,##0", "#,##0"]
        for ci, (val, nfmt) in enumerate(zip(values, number_fmts), start=2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font      = font(bold=(ci == 2), size=10)
            cell.fill      = fill(bg)
            cell.border    = border("thin")
            cell.alignment = align(h="left" if ci == 2 else "center", v="center")
            if nfmt:
                cell.number_format = nfmt
        r += 1

    trend_data_end = r - 1
    r = spacer(ws, r, 16)

    # ── 12-MONTH TREND CHART ──────────────────────────────────────────────────
    # Combo chart — dual y-axis:
    #   Left axis  (primary)   : clustered bars — New Subs (col C) + Renewals (col E)
    #   Right axis (secondary) : Trials line    (col D)
    try:
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.chart.legend import Legend
        from openpyxl.chart.marker import Marker
        from openpyxl.chart.label import DataLabelList

        bar = BarChart()
        bar.type     = "col"
        bar.grouping = "clustered"
        bar.title    = None          # no title
        bar.y_axis.title = "Subscriptions"
        bar.x_axis.title = None      # remove overlapping x-axis title
        bar.y_axis.axId  = 100
        bar.x_axis.axId  = 10

        # Categories: Month label (col B)
        cats = Reference(ws, min_col=2, min_row=trend_data_start, max_row=trend_data_end)
        bar.set_categories(cats)

        # Bar 1 — New Subs (col C) — light blue
        new_ref = Reference(ws, min_col=3, min_row=trend_hdr_row, max_row=trend_data_end)
        bar.add_data(new_ref, titles_from_data=True)
        bar.series[0].graphicalProperties.solidFill = "AED6F1"   # light blue

        # Bar 2 — Renewals (col E) — light teal
        renewal_ref = Reference(ws, min_col=5, min_row=trend_hdr_row, max_row=trend_data_end)
        bar.add_data(renewal_ref, titles_from_data=True)
        bar.series[1].graphicalProperties.solidFill = "A9DFBF"   # light green

        # Data labels on bars — dark font
        for s in bar.series:
            s.dLbls = DataLabelList()
            s.dLbls.showVal          = True
            s.dLbls.showLegendKey    = False
            s.dLbls.showCatName      = False
            s.dLbls.showSerName      = False

        # Line 1 — Paid New Subs (col K) on secondary (right) y-axis
        line_paid = LineChart()
        line_paid.y_axis.axId    = 300
        line_paid.y_axis.title   = "Paid New Subs / Trials"
        line_paid.y_axis.crosses = "max"
        paid_new_ref = Reference(ws, min_col=11, min_row=trend_hdr_row, max_row=trend_data_end)
        line_paid.add_data(paid_new_ref, titles_from_data=True)
        line_paid.series[0].graphicalProperties.line.solidFill = "1F3864"   # dark navy
        line_paid.series[0].graphicalProperties.line.width     = 20000
        line_paid.series[0].marker = Marker(symbol="circle", size=5)
        line_paid.series[0].dLbls = DataLabelList()
        line_paid.series[0].dLbls.showVal = True
        line_paid.series[0].dLbls.showLegendKey = False
        line_paid.series[0].dLbls.showCatName   = False
        line_paid.series[0].dLbls.showSerName   = False

        # Line 2 — Trials (col D) on secondary (right) y-axis
        line_trial = LineChart()
        line_trial.y_axis.axId    = 200
        line_trial.y_axis.crosses = "max"
        trial_ref = Reference(ws, min_col=4, min_row=trend_hdr_row, max_row=trend_data_end)
        line_trial.add_data(trial_ref, titles_from_data=True)
        line_trial.series[0].graphicalProperties.line.solidFill = "E74C3C"   # coral red
        line_trial.series[0].graphicalProperties.line.width     = 20000
        line_trial.series[0].marker = Marker(symbol="diamond", size=5)
        line_trial.series[0].dLbls = DataLabelList()
        line_trial.series[0].dLbls.showVal = True
        line_trial.series[0].dLbls.showLegendKey = False
        line_trial.series[0].dLbls.showCatName   = False
        line_trial.series[0].dLbls.showSerName   = False

        # Combine: bars (primary) + paid new line + trial line (secondary)
        bar += line_paid
        bar += line_trial

        # Remove all gridlines (must be set after combining)
        bar.y_axis.majorGridlines   = None
        bar.y_axis.minorGridlines   = None
        line_paid.y_axis.majorGridlines = None
        line_paid.y_axis.minorGridlines = None
        line_trial.y_axis.majorGridlines = None
        line_trial.y_axis.minorGridlines = None

        # Legend below chart
        bar.legend = Legend()
        bar.legend.position = "b"
        bar.legend.overlay  = False

        bar.width  = 28
        bar.height = 16

        ws.add_chart(bar, f"B{r}")
        r += 24
    except Exception as _chart_err:
        import logging
        logging.getLogger(__name__).warning(f"12-month chart skipped: {_chart_err}")


def _compute_var(current, previous) -> dict:
    """Compute variance dict for two values."""
    if previous is None or previous == 0:
              return None
    abs_var = current - previous
    pct_var = abs_var / abs(previous) * 100
    direction = "up" if abs_var > 0 else ("down" if abs_var < 0 else "flat")
    return {"abs": round(abs_var, 2), "pct": round(pct_var, 1), "direction": direction}
